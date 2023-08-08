import pandas as pd
import os
import geopandas as gpd
import shutil
import warnings
import wvdot_utils
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import datetime
warnings.filterwarnings("ignore")


class Cross_Validation():

    cols_dict = {
        'TRUCK':'NN',
    }

    def __init__(self, df):
        # self.df = load_defaults(df)
        self.df = df.rename(columns=self.cols_dict)
        try:
            self.df['BEGIN_DATE']
        except KeyError:
            self.df['BEGIN_DATE'] = datetime.datetime(2022,1,1)

    def check_geom(self,row):
        value = wvdot_utils.validate_geom.check_seg_valid(row['RouteID'],row['BMP'],row['EMP'])
        return value[-1]
    
    def sji01(self):
        #F_SYSTEM,F_SYSTEM ValueNumeric must not be NULL AND For every F_SYSTEM record; there should be a corresponding route on ARNOLD 
        print('Running rule SJI01...')
        self.df['SJI01'] = True
        self.df['SJI01_GEOM'] = self.df.apply(self.check_geom,axis=1)
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM'].notna() & ~(tmpDF['SJI01_GEOM'])]
        self.df.drop('SJI01_GEOM',axis=1,inplace = True)
        # tmpDF = tmpDF[tmpDF.apply(cross_validation.check_geom,axis=1)]
        self.df['SJI01'].iloc[tmpDF.index.tolist()] = False


    def sji02(self):
        #FACILITY_TYPE ValueNumeric must not be null AND For every FACILITY_TYPE record; there should be a corresponding route on ARNOLD
        print('Running rule SJI02...')
        self.df['SJI02'] = True
        self.df['SJI02_GEOM'] = self.df.apply(self.check_geom,axis= 1)
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE'].notna() & ~(tmpDF['SJI02_GEOM'])]
        self.df.drop('SJI02_GEOM',axis=1,inplace=True)
        self.df['SJI02'].iloc[tmpDF.index.tolist()] =False


    def sji03(self):
        # OWNERHIP ValueNumeric must not be NULL AND For every OWNERSHIP record; there should be a corresponding route on ARNOLD
        print('Running rule SJI03...')
        self.df['SJI03'] = True
        self.df['SJI03_GEOM'] = self.df.apply(self.check_geom,axis= 1)
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['OWNERSHIP'].notna() & ~(tmpDF['SJI03_GEOM'])]
        self.df.drop('SJI03_GEOM',axis = 1, inplace = True)
        self.df['SJI03'].iloc[tmpDF.index.tolist()] = False

    def sji04(self):
        # URBAN_ID ValueNumeric must not be NULL AND For every URBAN_ID record; there should be a corresponding route on ARNOLD
        print('Running rule SJI04...')
        self.df['SJI04'] = True
        self.df['SJI04_GEOM'] = self.df.apply(self.check_geom,axis = 1)
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['URBAN_CODE'].notna() & ~(tmpDF['SJI04_GEOM'])]
        self.df.drop('SJI04_GEOM',axis = 1, inplace = True)
        self.df['SJI04'].iloc[tmpDF.index.tolist()] = False

    def sji05(self):
        # F_SYSTEM must exist where FACILITY_TYPE exists
        print('Running rule SJI05...')
        self.df['SJI05'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE'].notna()]
        tmpDF = tmpDF[tmpDF['F_SYSTEM'].isna()]
        self.df['SJI05'].iloc[tmpDF.index.tolist()] = False

    def sji06(self):
        # FACILITY_TYPE must exist where F_SYSTEM exists
        print('Running rule SJI06...')
        self.df['SJI06'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM'].notna()]
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE'].isna()]
        self.df['SJI06'].iloc[tmpDF.index.tolist()] = False

    def sji07(self):
        # Where F_SYSTEM ValueNumeric = 1; FACILITY_TYPE ValueNumeric in (1;2) must exist and ROUTE_NUMBER ValueNumeric or ValueText must not be NULL
        print('Running rule SJI07...')
        self.df['SJI07'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM'] == 1]
        tmpDF = tmpDF[~tmpDF['FACILITY_TYPE'].isin([1,2]) | (tmpDF['ROUTE_NUMBER'].isna() & tmpDF['ROUTE_NUMBER_VALUE_TEXT'].isna())]
        self.df['SJI07'].iloc[tmpDF.index.tolist()] = False

    def sji08(self):
        # If F_SYSTEM ValueNumeric = 1 AND FACILITY_TYPE is IN (1,2) then NHS must exist and NHS ValueNumeric must = 1
        print('Running rule SJI08...')
        self.df['SJI08'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM']==1]
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE'].isin([1,2])]
        tmpDF = tmpDF[tmpDF['NHS'].isna() | (tmpDF['NHS']!=1)]
        self.df['SJI08'].iloc[tmpDF.index.tolist()] = False

    def sji09(self):
        # Where ROUTE_NUMBER ValueNumeric or ValueText is not NULL; ROUTE_SIGNING ValueNumeric must not be NULL
        print('Runing rule SJI09...')
        self.df['SJI09'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['ROUTE_NUMBER'].notna() | (tmpDF['ROUTE_NUMBER_VALUE_TEXT'].notna()) ]
        tmpDF = tmpDF[tmpDF['ROUTE_SIGNING'].isna()]
        self.df['SJI09'].iloc[tmpDF.index.tolist()] = False

    def sji10(self):
        # Where ROUTE_NUMBER ValueNumeric or ValueText is not NULL; ROUTE_QUALIFIER ValueNumeric must not be NULL
        print('Running rule SJI10...')
        self.df['SJI10'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['ROUTE_NUMBER'].notna() | (tmpDF['ROUTE_NUMBER_VALUE_TEXT'].notna())]
        tmpDF = tmpDF[tmpDF['ROUTE_QUALIFIER'].isna()]
        self.df['SJI10'].iloc[tmpDF.index.tolist()] = False

    def sji11(self):
        # If F_SYSTEM ValueNumeric = 1 Then STRAHNET_TYPE must exist and STRAHNET_TYPE ValueNumeric must = 1
        print('Running rule SJI11...')
        self.df['SJI11'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM']==1]
        tmpDF = tmpDF[tmpDF['STRAHNET_TYPE'].isna() | (tmpDF['STRAHNET_TYPE']!=1)]
        self.df['SJI11'].iloc[tmpDF.index.tolist()] = False
    
    def sji12(self):
        # If STRAHNET_TYPE ValueNumeric is in the range (1;2) then NHS ValueNumeric must = 1
        print('Running rule SJI12...')
        self.df['SJI12'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['STRAHNET_TYPE'].isin([1,2])]
        tmpDF = tmpDF[tmpDF['NHS']!=1]
        self.df['SJI12'].iloc[tmpDF.index.tolist()] = False

    def sji13(self):
        # If F_SYSTEM ValueNumeric = 1 Then NN must exist and NN ValueNumeric must = 1
        print('Running rule SJI13')
        self.df['SJI13'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['F_SYSTEM']==1]
        tmpDF = tmpDF[tmpDF['NN'].isna() | (tmpDF['NN']!=1)]
        self.df['SJI13'].iloc[tmpDF.index.tolist()] = False

    def sjt01(self):
        # AADT_SINGLE_UNIT should be < (AADT*0.4)
        print('Running rule SJT01...')
        self.df['SJT01'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['AADT_SINGLE_UNIT'].notna()]
        tmpDF = tmpDF[tmpDF['AADT_SINGLE_UNIT'] > (tmpDF['AADT']*0.4)]
        self.df['SJT01'].iloc[tmpDF.index.tolist()] = False

    def sjt02(self):
        # IF AADT > 500 THEN AADT_SINGLE_UNIT should be > 0
        print('Running rule SJT02...')
        self.df['SJT02'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['AADT']>500]
        tmpDF = tmpDF[tmpDF['AADT_SINGLE_UNIT']<0]
        self.df['SJT02'].iloc[tmpDF.index.tolist()] = False

    def sjt03(self):
        # AADT_SINGLE_UNIT + AADT_COMBINATION should be < (0.8*AADT)
        print('Running rule SJT03...')
        self.df['SJT03'] = True
        tmpDF = self.df.copy()
        tmpDF['AADT_SINGLE_UNIT'].fillna(0, inplace=True)
        tmpDF['AADT_COMBINATION'].fillna(0, inplace=True)
        tmpDF['SJT03_SUM'] = tmpDF['AADT_SINGLE_UNIT'] + tmpDF['AADT_COMBINATION']
        tmpDF = tmpDF[tmpDF['SJT03_SUM'] >= (tmpDF['AADT']*0.8)]
        self.df['SJT03'].iloc[tmpDF.index.tolist()] = False

    def sjt04(self):
        # (AADT_SINGLE_UNIT x 0.01) < (AADT x (PCT_DH_SINGLE_UNIT/100)) < (AADT_SINGLE_UNIT x 0.5)  
        print('Running rule SJT04...')
        self.df['SJT04'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[((tmpDF['AADT_SINGLE_UNIT']*0.01) >= (tmpDF['AADT']*(tmpDF['PCT_DH_SINGLE_UNIT']/100)))|((tmpDF['AADT']*(tmpDF['PCT_DH_SINGLE_UNIT']/100))>=(tmpDF['AADT_SINGLE_UNIT']*0.5))]
        self.df['SJT04'].iloc[tmpDF.index.tolist()] = False
    
    def sjt05(self):
        # PCT_DH_SINGLE_UNIT should be > 0 and < 25%
        print('Running rule SJT05...')
        self.df['SJT05'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[(tmpDF['PCT_DH_SINGLE_UNIT'] <= 0) | (tmpDF['PCT_DH_SINGLE_UNIT'] >= 25) ]
        self.df['SJT05'].iloc[tmpDF.index.tolist()] = False
    
    def sjt06(self):
        # AADT_SINGLE_UNIT should be < 50 Where PCT_DH_SINGLE_UNIT = 0
        print('Running Rule SJT06...')
        self.df['SJT06'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['PCT_DH_SINGLE_UNIT']==0]
        tmpDF = tmpDF[tmpDF['AADT_SINGLE_UNIT']>=50]
        self.df['SJT06'].iloc[tmpDF.index.tolist()] = False

    def sjt07(self):
        # AADT_COMBINATION should be < (AADT*0.4)
        print('Running rule SJT07...')
        self.df['SJT07'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['AADT_COMBINATION'].notna()]
        tmpDF = tmpDF[tmpDF['AADT_COMBINATION'] > (tmpDF['AADT']*0.4)]
        self.df['SJT07'].iloc[tmpDF.index.tolist()] = False

    def sjt08(self):
        # If AADT is > 500 then AADT_COMBINATION Should be > 0 
        print('Running rule SJT08...')
        self.df['SJT08'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['AADT']>500]
        tmpDF = tmpDF[tmpDF['AADT_COMBINATION']<0]
        self.df['SJT08'].iloc[tmpDF.index.tolist()] = False

    def sjt09(self):
        # (AADT_COMBINATION x 0.01) < (AADT x (PCT_DH_COMBINATION/100)) < (AADT_COMBINATION x 0.5)  
        print('Running rule SJT09...')
        self.df['SJT09'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[((tmpDF['AADT_COMBINATION']*0.01) >= (tmpDF['AADT']*(tmpDF['PCT_DH_COMBINATION']/100)))|((tmpDF['AADT']*(tmpDF['PCT_DH_COMBINATION']/100))>=(tmpDF['AADT_COMBINATION']*0.5))]
        self.df['SJT04'].iloc[tmpDF.index.tolist()] = False

    def sjt10(self):
        # PCT_DH_COMBINATION should be > 0 and < 25%
        print('Running rule SJT10...')
        self.df['SJT10'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[(tmpDF['PCT_DH_COMBINATION'] <= 0) | (tmpDF['PCT_DH_COMBINATION'] >= 25) ]
        self.df['SJT10'].iloc[tmpDF.index.tolist()] = False

    def sjt11(self):
        # AADT_COMBINATION should be < 50 Where PCT_DH_COMBINATION = 0
        print('Running rule SJT11...')
        self.df['SJT11'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['PCT_DH_COMBINATION']==0]
        tmpDF = tmpDF[tmpDF['AADT_COMBINATION']>=50]
        self.df['SJT11'].iloc[tmpDF.index.tolist()] = False

    def sjt12(self):
        # K_FACTOR ValueNumeric MUST BE > 4 and <30
        print('Running rule SJT12...')
        self.df['SJT12'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[(tmpDF['K_FACTOR']<=0) | (tmpDF['K_FACTOR'] >=30)]
        self.df['SJT12'].iloc[tmpDF.index.tolist()] = False

    def sjt13(self):
        # DIR_FACTOR ValueNumeric must =100 where FACILITY_TYPE ValueNumeric = 1
        print('Running rule SJT13...')
        self.df['SJT13'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE']==1]
        tmpDF = tmpDF[tmpDF['DIR_FACTOR']!=100]
        self.df['SJT13'].iloc[tmpDF.index.tolist()] = False

    def sjt14(self):
        # DIR_FACTOR ValueNumeric must be >50 and <=75 where FACILITY_TYPE ValueNumeric = 2
        print('Running rule SJT14...')
        self.df['SJT14'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['FACILITY_TYPE']==2]
        tmpDF = tmpDF[(tmpDF['DIR_FACTOR']<=50) | (tmpDF['DIR_FACTOR']>=75) ]
        self.df['SJT14'].iloc[tmpDF.index.tolist()] = False

    def sjt15(self):
        # Future AADT must be > AADT and < 4* AADT where ValueDate is Null; OR Future AADT must be < AADT * 0.2(ValueDate-BeginDate year)
        print('Running rule SJT15...')
        self.df['SJT15'] = True
        tmpDF = self.df.copy()
        tmpDF = tmpDF[tmpDF['FUTURE_AADT_VALUE_DATE'].isna()]
        tmpDF = tmpDF[(tmpDF['FUTURE_AADT']<= tmpDF['AADT']) | (tmpDF['FUTURE_AADT']>= (4*tmpDF['AADT']))]
        self.df['SJT15'].iloc[tmpDF.index.tolist()] = False

        tmpDF = self.df.copy()
        tmpDF['FUTURE_AADT_VALUE_DATE'] = pd.to_datetime(tmpDF['FUTURE_AADT_VALUE_DATE'])
        tmpDF = tmpDF[tmpDF['FUTURE_AADT_VALUE_DATE'].notna()]
        tmpDF = tmpDF[(tmpDF['FUTURE_AADT'])>= (tmpDF['AADT']*0.2*(tmpDF['FUTURE_AADT_VALUE_DATE'].dt.year-tmpDF['BEGIN_DATE'].dt.year))]
        self.df['SJT15'].iloc[tmpDF.index.tolist()] = False



    def run(self,inventory=True,traffic=True):
        if inventory==True:
            print('Running Spatial Join Inventory')
            self.sji01()
            self.sji02()
            self.sji03()
            self.sji04()
            self.sji05()
            self.sji06()
            self.sji07()
            self.sji08()
            self.sji09()
            self.sji10()
            self.sji11()
            self.sji12()
            self.sji13()
        if traffic==True:
            print('Running Spatial Join Traffic')    
            self.sjt01()
            self.sjt02()
            self.sjt03()
            self.sjt04()
            self.sjt05()
            self.sjt06()
            self.sjt07()
            self.sjt08()
            self.sjt09()
            self.sjt10()
            self.sjt11()
            self.sjt12()
            self.sjt13()
            self.sjt14()
            self.sjt15()





    def create_output(self, template='templates/cross_validation_rules_template.xlsx', outfilename='summary/cross_validation_rules_summary.xlsx'):
        #Reads sheet on template that list all data items associated with each rule and converts to dictionary
        dataItemsDF = pd.read_excel(template, sheet_name="ruleDataItems", usecols='A,B', nrows=28)
        dataItemsDF['Rule'] = dataItemsDF['Rule'].str.replace("-", "")
        dataItemsDF['Data_Items'] = dataItemsDF['Data_Items'].str.split(",")
        ruleDict = dict(zip(dataItemsDF['Rule'], dataItemsDF['Data_Items']))
        fwhaRuleDict = ruleDict.copy()
        for rule in ruleDict.keys():
            fwhaRuleDict[rule+"_FHWA"] = fwhaRuleDict.pop(rule)


        #Reads the rule descripts off of summary sheet and converts to dictionary
        ruleDescDF = pd.read_excel(template, sheet_name="Summary", usecols="A,D")
        ruleDescDF['Rule'] = ruleDescDF['Rule'].str.replace("-", "")
        ruleDesc = dict(zip(ruleDescDF['Rule'], ruleDescDF['Description']))

        #Create copy of template to write to
        shutil.copy(template, outfilename)

        with pd.ExcelWriter(outfilename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            tempDF = self.df.copy()
            numFailed = []
            numPassed = []
            lenFailed = []
            fwhaFailed = []

            #Get counts for failed/passed/length of failed
            print("Updating summary sheet on",outfilename,"...")
            for rule in ruleDict.keys():
                #Assumes all passes for rules not ran
                try:
                    numFailed.append(tempDF[tempDF[rule]==False].shape[0])
                    numPassed.append(tempDF[tempDF[rule]==True].shape[0])
                    lenFailed.append((tempDF[tempDF[rule]==False]['EMP'] - tempDF[tempDF[rule]==False]['BMP']).sum())
                except KeyError:
                    numFailed.append(0)
                    numPassed.append(self.df.shape[0])
                    lenFailed.append(0)

            for rule in fwhaRuleDict.keys():
                try:
                    fwhaFailed.append(tempDF[tempDF[rule]==False].shape[0])
                except KeyError:
                    fwhaFailed.append(0)

            failedDF = pd.DataFrame(numFailed)
            passedDF = pd.DataFrame(numPassed)
            lengthDF = pd.DataFrame(lenFailed)
            fwhaFailedDF = pd.DataFrame(fwhaFailed)

            #Write counts to Summary sheet
            failedDF.to_excel(writer, sheet_name='Summary', startcol=4, startrow=1, header=False, index=False)
            passedDF.to_excel(writer, sheet_name='Summary', startcol=5, startrow=1, header=False, index=False)
            lengthDF.to_excel(writer, sheet_name='Summary', startcol=6, startrow=1, header=False, index=False)
            fwhaFailedDF.to_excel(writer, sheet_name='Summary', startcol=7, startrow=1, header=False, index=False)


            #Create sheets for each rule containing all failed rows (using only columns that the specific rule references)
            for rule in ruleDict.keys():
                tempDF = self.df.copy()
                #Checks to make sure rule has data items associated with it (will be a list if dataItems exists, otherwise will be float (np.nan))
                if type(ruleDict[rule])==list:
                    #Tries using RULENAME (i.e. SJF01) in dataset which is added if the rule is ran
                    #If rule is not run, no column will exist with the rulename, catches KeyError and prints message.
                    try:
                        if tempDF[tempDF[rule]==False].shape[0] > 0:
                            print("Creating error sheet for rule:",rule)
                            dataItems = ['RouteID', 'BMP', 'EMP']
                            [dataItems.append(x) for x in ruleDict[rule] if x not in dataItems]
                            tempDF = tempDF[tempDF[rule]==False]
                            tempDF = tempDF[dataItems]
                            tempDF.to_excel(writer, sheet_name=rule, startrow=1, index=False)
                            worksheet = writer.sheets[rule]
                            worksheet['A1'] = f'=HYPERLINK("#Summary!A1", "Summary Worksheet")'
                            worksheet['A1'].font = Font(underline='single', color='0000EE')
                            #Autofit columns
                            for i in range(1, worksheet.max_column+1):
                                worksheet.column_dimensions[get_column_letter(i)].width = 20
                            #Add rule description to sheet
                            worksheet['B1'] = ruleDesc[rule]
                            

                        else:
                            print("No failed rows for rule:",rule)

                    except KeyError:
                        print(rule,"not found in DF. Sheet was not created in rules_summary.xlsx")
                else:
                    print("No data items for rule",rule,", Sheet not created.")



# input_file = 'all_submission_data.csv'
# data = pd.read_csv(input_file)
# cross_validation = Cross_Validation(data)
# cross_validation.run(3)
# cross_validation.create_output()
