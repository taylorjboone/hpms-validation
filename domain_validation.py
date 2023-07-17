import pandas as pd
import shutil
import os
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import warnings
from datetime import datetime
import numpy as np
warnings.filterwarnings("ignore")

URBAN_CODE_list = ['06139','15481','21745','36190','40753','59275','67672','93592','94726', '99998', '99999']
#Matches dates in format MM/DD/YYYY
date_regex = "(0[1-9]|1[012])/(0[1-9]|[12][0-9]|3[01])/(19|20)\d\d"

class domain_validations():
    
    def __init__(self,df):
        self.df = df
        self.df.rename(columns={'TRUCK':'NN'}, inplace=True)

        try:
            self.df['BEGIN_DATE']
        except:
            self.df['BEGIN_DATE'] = datetime(2022,1,1)

    def drd01(self):
        #F_SYSTEM ValueNumeric must be a valid integer in the range (1-7)
        print("Running rule DRD01...")
        self.df['DRD01'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].notna()]
        tempDF = tempDF[~tempDF['F_SYSTEM'].isin(range(1,8))]
        self.df['DRD01'].iloc[tempDF.index.tolist()] = False

    def drd64(self):
        #NHS ValueNumeric must be an integer in the range (1-9)

        print("Running rule DRD64...")
        self.df['DRD64'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].notna()]
        tempDF = tempDF[~tempDF['NHS'].isin(range(1,10))]
        self.df['DRD64'].iloc[tempDF.index.tolist()] = False

    def drd65(self):
        #STRAHNET_TYPE ValueNumeric must be an integer in the range (1;2)

        print("Running rule DRD65...")
        self.df['DRD65'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['STRAHNET_TYPE'].notna()]
        tempDF = tempDF[~tempDF['STRAHNET_TYPE'].isin([1,2])]
        self.df['DRD65'].iloc[tempDF.index.tolist()] = False

    def drd66(self):
        #NN ValueNumeric must be an integer in the range (1;2)

        print("Running rule DRD66...")
        self.df['DRD66'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NN'].notna()]
        tempDF = tempDF[~tempDF['NN'].isin([1,2])]
        self.df['DRD66'].iloc[tempDF.index.tolist()] = False

    def drd72(self):
        #NHFN ValueNumeric must be an integer in the range (1;2;3)
        #NHFN isn't reported, so rule is disabled in the list below
        print("Running rule DRD72...")
        self.df['DRD72'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHFN'].notna()]
        tempDF = tempDF[~tempDF['NHFN'].isin([1,2,3])]
        self.df['DRD72'].iloc[tempDF.index.tolist()] = False


    def drd202(self):
        #RouteID field must not be blank and can not contain pipe characters 
        print("Running rule DRD202...")
        self.df['DRD202'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RouteID'].isna() | tempDF['RouteID'].str.contains("|", regex=False, na=False)]
        self.df['DRD202'].iloc[tempDF.index.tolist()] = False

    def drd203(self):
        #BeginPoint must be in the format Numeric (8;3) AND must be < EndPoint
        print("Running rule DRD203...")
        self.df['DRD203'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['BMP'] > tempDF['EMP']]
        self.df['DRD203'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF['BMP'] = tempDF['BMP'].astype("string")
        tempDF['BMP'] = tempDF['BMP'].str.split(".")
        tempDF = tempDF[(tempDF['BMP'].str[0].str.len() > 8) | (tempDF['BMP'].str[1].str.len() > 3)]
        self.df['DRD203'].iloc[tempDF.index.tolist()] = False

    def drd204(self):
        #EndPoint must be in the format Numeric (8;3) AND must be > BeginPoint

        print("Running rule DRD204...")
        self.df['DRD204'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['EMP'] < tempDF['BMP']]
        self.df['DRD204'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF['EMP'] = tempDF['EMP'].astype("string")
        tempDF['EMP'] = tempDF['EMP'].str.split(".")
        tempDF = tempDF[(tempDF['EMP'].str[0].str.len() > 8) | (tempDF['EMP'].str[1].str.len() > 3)]
        self.df['DRD204'].iloc[tempDF.index.tolist()] = False

    def dre02(self):
        #URBAN_ID ValueNumeric must be an integer and a valid US Census Urban Code for the applicable state

        print("Running rule DRE02...")
        self.df['DRE02'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['URBAN_CODE'].notna()]
        tempDF = tempDF[~tempDF['URBAN_CODE'].isin(URBAN_CODE_list) | tempDF['URBAN_CODE'].str.contains('.', regex=False)]
        self.df['DRE02'].iloc[tempDF.index.tolist()] = False  

    def dre03(self):
        #FACILITY_TYPE ValueNumeric must be a valid integer in the range (1;2;4;5;6;7)

        print("Running rule DRE03...")
        self.df['DRE03'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].notna()]
        tempDF = tempDF[~tempDF['FACILITY_TYPE'].isin([1,2,4,5,6,7])]
        self.df['DRE03'].iloc[tempDF.index.tolist()] = False

    def dre04(self):
        #STRUCTURE_TYPE ValueNumeric must be a valid integer within the range (1-3)

        print("Running rule DRE04...")
        self.df['DRE04'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'].notna()]
        tempDF = tempDF[~tempDF['STRUCTURE_TYPE'].isin([1,2,3])]
        self.df['DRE04'].iloc[tempDF.index.tolist()] = False

    def dre05(self):
        #ACCESS_CONTROL ValueNumeric must be a valid integer within the range (1-3)

        print("Running rule DRE05...")
        self.df['DRE05'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['ACCESS_CONTROL'].notna()]
        tempDF = tempDF[~tempDF['ACCESS_CONTROL'].isin([1,2,3])]
        self.df['DRE05'].iloc[tempDF.index.tolist()] = False

    def dre06(self):
        #OWNERSHIP ValueNumeric must be a valid integer within the range (1;2;3;4;11;12;21;25;26;27;31;32;40;50;60;62;63;64;66;67;68;69;70;72;73;74;80)

        print("Running rule DRE06...")
        self.df['DRE06'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['OWNERSHIP'].notna()]
        tempDF = tempDF[~tempDF['OWNERSHIP'].isin([1,2,3,4,11,12,21,25,26,27,31,32,40,50,60,62,63,64,66,67,68,69,70,72,73,74,80])]
        self.df['DRE06'].iloc[tempDF.index.tolist()] = False
                
    def dre07(self):
        #THROUGH_LANES ValueNumeric must be a valid integer > 0 

        print("Running rule DRE07...")
        self.df['DRE07'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['THROUGH_LANES'].notna()]
        tempDF = tempDF[(tempDF['THROUGH_LANES'] <= 0) | (tempDF['THROUGH_LANES']%np.floor(tempDF['THROUGH_LANES']) != 0)]
        self.df['DRE07'].iloc[tempDF.index.tolist()] = False

    def dre10(self):
        #PEAK_LANES ValueNumeric must be a valid integer > 0

        print("Running rule DRE10...")
        self.df['DRE10'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PEAK_LANES'].notna()]
        tempDF = tempDF[tempDF['PEAK_LANES'] <= 0]
        self.df['DRE10'].iloc[tempDF.index.tolist()] = False

    def dre11(self):
        #COUNTER_PEAK_LANES ValueNumeric must be a valid integer > 0 

        print("Running rule DRE11...")
        self.df['DRE11'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['COUNTER_PEAK_LANES'].notna()]
        tempDF = tempDF[tempDF['COUNTER_PEAK_LANES'] <= 0]
        self.df['DRE11'].iloc[tempDF.index.tolist()] = False

    def dre12(self):
        #TURN_LANES_R ValueNumeric must be a valid integer within the range (1-6)

        print("Running rule DRE12...")
        self.df['DRE12'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['TURN_LANES_R'].notna()]
        tempDF = tempDF[~tempDF['TURN_LANES_R'].isin(range(1,7))]
        self.df['DRE12'].iloc[tempDF.index.tolist()] = False

    def dre13(self):
        #TURN_LANES_L ValueNumeric must be a valid integer within the range (1-6)

        print("Running rule DRE13...")
        self.df['DRE13'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['TURN_LANES_L'].notna()]
        tempDF = tempDF[~tempDF['TURN_LANES_L'].isin(range(1,7))]
        self.df['DRE13'].iloc[tempDF.index.tolist()] = False

    def dre14(self):
        #SPEED_LIMIT ValueNumeric must be a valid integer in the range (0-99)

        print("Running rule DRE14...")
        self.df['DRE14'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SPEED_LIMIT'].notna()]
        tempDF = tempDF[~tempDF['SPEED_LIMIT'].isin(range(0,100))]
        self.df['DRE14'].iloc[tempDF.index.tolist()] = False

    def dre15(self):
        #TOLL_ID ValueNumeric must be a valid Toll ID of no more than 15
        #Assuming this rule means no more than 15 characters, as most toll IDs are > 15

        print("Running rule DRE15...")
        self.df['DRE15'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['TOLL_ID'].notna()]
        tempDF = tempDF[tempDF['TOLL_ID'].astype("string").str.len() > 15]
        self.df['DRE15'].iloc[tempDF.index.tolist()] = False

    def dre21(self):
        #AADT ValueNumeric must be positive integer in the range (0-600000); 
        # ValueText must be in (A;B;C;D;E); 
        # and ValueDate must not be null; and must be in the format YYYY

        #Writing the date check as must be in the format MM/DD/YYYY as that's current working on the website. Validation rule
        #and HPMS manual disagree on this on
        print("Running rule DRE21...")
        self.df['DRE21'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT'].notna()]
        tempDF = tempDF[~tempDF['AADT'].isin(range(0,600001))]
        self.df['DRE21'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_VALUE_TEXT'].notna()]
        tempDF = tempDF[~tempDF['AADT_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE21'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT'].notna()]
        tempDF = tempDF[tempDF['AADT_VALUE_DATE'].isna()]
        self.df['DRE21'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_VALUE_DATE'].notna()]
        tempDF = tempDF[~tempDF['AADT_VALUE_DATE'].str.contains(date_regex)]
        self.df['DRE21'].iloc[tempDF.index.tolist()] = False

    def dre22(self):
        #AADT_SINGLE_UNIT ValueNumeric must be positive integer in the range (0-500000);
        #  ValueText must be in (A;B;C;D;E)

        print("Running rule DRE22...")
        self.df['DRE22'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_SINGLE_UNIT'].notna()]
        tempDF = tempDF[~tempDF['AADT_SINGLE_UNIT'].isin(range(0,500001))]
        self.df['DRE22'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_SINGLE_UNIT_VALUE_TEXT'].notna()]
        tempDF = tempDF[~tempDF['AADT_SINGLE_UNIT_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE22'].iloc[tempDF.index.tolist()] = False

    def dre23(self):
        #PCT_DH_SINGLE_UNIT ValueNumeric must be in the range (0-50)

        print("Running rule DRE23...")
        self.df['DRE23'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PCT_DH_SINGLE_UNIT'].notna()]
        tempDF = tempDF[(tempDF['PCT_DH_SINGLE_UNIT'] > 50) | (tempDF['PCT_DH_SINGLE_UNIT'] < 0)]
        self.df['DRE23'].iloc[tempDF.index.tolist()] = False   

    def dre24(self):
        #AADT_COMBINATION ValueNumeric must be positive integer in the range (0-500000); ValueText must be in (A;B;C;D;E)

        print("Running rule DRE24...")
        self.df['DRE24'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_COMBINATION'].notna()]
        tempDF = tempDF[~tempDF['AADT_COMBINATION'].isin(range(0,500001))]
        self.df['DRE24'].iloc[tempDF.index.tolist()] = False   

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AADT_COMBINATION_VALUE_TEXT'].notna()]
        tempDF = tempDF[~tempDF['AADT_COMBINATION_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE24'].iloc[tempDF.index.tolist()] = False 

    def dre25(self):
        #PCT_DH_COMBINATION ValueNumeric must be in the range (0-50)
        print("Running rule DRE25...")
        self.df['DRE25'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PCT_DH_COMBINATION'].notna()]
        tempDF = tempDF[(tempDF['PCT_DH_COMBINATION'] < 0) | (tempDF['PCT_DH_COMBINATION'] > 50)]
        self.df['DRE25'].iloc[tempDF.index.tolist()] = False   

    def dre26(self):
        #K_FACTOR ValueNumeric must be a positive integer > 4

        print("Running rule DRE26...")
        self.df['DRE26'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['K_FACTOR'] <= 4]
        self.df['DRE26'].iloc[tempDF.index.tolist()] = False

    def dre27(self):
        #DIR_FACTOR ValueNumeric must be a postitive integer in the range (50-100)

        print("Running rule DRE27...")
        self.df['DRE27'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['DIR_FACTOR'].notna()]
        tempDF = tempDF[~tempDF['DIR_FACTOR'].isin(range(50,101))]
        self.df['DRE27'].iloc[tempDF.index.tolist()] = False

    def dre28(self):
        #FUTURE_AADT ValueNumeric must be a positive integer AND ValueDate must be >= BeginDate + 18 AND <= BeginDate +25

        print("Running rule DRE28...")
        self.df['DRE28'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FUTURE_AADT'].notna()]
        tempDF = tempDF[(tempDF['FUTURE_AADT'] <= 0) | (tempDF['FUTURE_AADT']%np.floor(tempDF['FUTURE_AADT']) != 0)]
        self.df['DRE28'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF['FUTURE_AADT_VALUE_DATE'] = pd.to_datetime(tempDF['FUTURE_AADT_VALUE_DATE'])
        tempDF = tempDF[(tempDF['FUTURE_AADT_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year + 18)) | (tempDF['FUTURE_AADT_VALUE_DATE'].dt.year > (tempDF['BEGIN_DATE'].dt.year + 25))]
        self.df['DRE28'].iloc[tempDF.index.tolist()] = False

    def dre29(self):
        #SIGNAL_Type ValueNumeric must be an integer in the range (1 -5)

        print('Running rule DRE29...')
        self.df['DRE29'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SIGNAL_TYPE'].notna()]
        tempDF = tempDF[~tempDF['SIGNAL_TYPE'].isin(range(1,6))]
        self.df['DRE29'].iloc[tempDF.index.tolist()] = False

    def dre30(self):
        #PCT_GREEN_TIME ValueNumeric must be a positive integer > 0 and <=100

        print('Running rule DRE30...')
        self.df['DRE30'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PCT_GREEN_TIME'].notna()]
        tempDF = tempDF[~tempDF['PCT_GREEN_TIME'].isin(range(1,101))]
        self.df['DRE30'].iloc[tempDF.index.tolist()] = False

    def dre31(self):
        #NUMBER_SIGNALS ValueNumeric must be an integer >=0

        print('Running rule DRE31...')
        self.df['DRE31'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NUMBER_SIGNALS'].notna()]
        tempDF = tempDF[tempDF['NUMBER_SIGNALS']!=0]
        tempDF = tempDF[(tempDF['NUMBER_SIGNALS'] < 0) | (tempDF['NUMBER_SIGNALS']%np.floor(tempDF['NUMBER_SIGNALS'])!=0)]
        self.df['DRE31'].iloc[tempDF.index.tolist()] = False

    def dre32(self):
        #STOP_SIGNS ValueNumeric must be an integer >=0

        print('Running rule DRE32...')
        self.df['DRE32'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['STOP_SIGNS'].notna()]
        tempDF = tempDF[tempDF['STOP_SIGNS']!=0]
        tempDF = tempDF[(tempDF['STOP_SIGNS'] < 0) | (tempDF['STOP_SIGNS']%np.floor(tempDF['STOP_SIGNS']) != 0)]
        self.df['DRE32'].iloc[tempDF.index.tolist()] = False

    def dre33(self):
        #AT_GRADE_OTHER ValueNumeric must be an integer >=0

        print('Running rule DRE33...')
        self.df['DRE33'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['AT_GRADE_OTHER'].notna()]
        tempDF = tempDF[tempDF['AT_GRADE_OTHER'] != 0]
        tempDF = tempDF[(tempDF['AT_GRADE_OTHER'] < 0) | (tempDF['AT_GRADE_OTHER']%np.floor(tempDF['AT_GRADE_OTHER'])!=0)]
        self.df['DRE33'].iloc[tempDF.index.tolist()] = False

    def dre34(self):
        #LANE_WIDTH ValueNumberic must be an integer in the range (6-30)

        print('Running rule DRE34...')
        self.df['DRE34'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['LANE_WIDTH'].notna()]
        tempDF = tempDF[~tempDF['LANE_WIDTH'].isin(range(6,31))]
        self.df['DRE34'].iloc[tempDF.index.tolist()] = False

    def dre35(self):
        #MEDIAN_TYPE ValueNumeric must be an integer in the range (1-7)

        print('Running rule DRE35...')
        self.df['DRE35'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].notna()]
        tempDF = tempDF[~tempDF['MEDIAN_TYPE'].isin(range(1,8))]
        self.df['DRE35'].iloc[tempDF.index.tolist()] = False

    def dre36(self):
        #MEDIAN_WIDTH ValueNumeric must be an integer in the range (0-99)

        print('Running rule DRE36...')
        self.df['DRE36'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['MEDIAN_WIDTH'].notna()]
        tempDF = tempDF[~tempDF['MEDIAN_WIDTH'].isin(range(0,100))]
        self.df['DRE36'].iloc[tempDF.index.tolist()] = False

    def dre37(self):
        #SHOULDER_TYPE ValueNumeric must be an integer in the range (1-6)

        print('Running rule DRE37...')
        self.df['DRE37'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SHOULDER_TYPE'].notna()]
        tempDF = tempDF[~tempDF['SHOULDER_TYPE'].isin(range(1,7))]
        self.df['DRE37'].iloc[tempDF.index.tolist()] = False

    def dre38(self):
        #SHOULDER_WIDTH_R ValueNumeric must be an integer >=0

        print('Running rule DRE38...')
        self.df['DRE38'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_R'].notna()]
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_R'] != 0]
        tempDF = tempDF[(tempDF['SHOULDER_WIDTH_R'] < 0) | (tempDF['SHOULDER_WIDTH_R']%np.floor(tempDF['SHOULDER_WIDTH_R']) != 0)]
        self.df['DRE38'].iloc[tempDF.index.tolist()] = False

    def dre39(self):
        #SHOULDER_WIDTH_L ValueNumeric must be an integer >=0

        print('Running rule DRE39...')
        self.df['DRE39'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_L'].notna()]
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_L'] != 0]
        tempDF = tempDF[(tempDF['SHOULDER_WIDTH_L'] < 0) | (tempDF['SHOULDER_WIDTH_L']%np.floor(tempDF['SHOULDER_WIDTH_L']) != 0)]
        self.df['DRE39'].iloc[tempDF.index.tolist()] = False

    def dre40(self):
        #PEAK_PARKING ValueNumeric must be an integer in the range (1-3)

        print('Running rule DRE40...')
        self.df['DRE40'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PEAK_PARKING'].notna()]
        tempDF = tempDF[~tempDF['PEAK_PARKING'].isin(range(1,4))]
        self.df['DRE40'].iloc[tempDF.index.tolist()] = False

    def dre42(self):
        #WIDENING_POTENTIAL ValueNumberic must be an integer in the range (1-4) AND ValueText must be (X) OR (A;B;C;D;E). 
        #Where ValueText = X; ValueNumeric Must be in (1-4). 

        print('Running rule DRE42...')
        self.df['DRE42'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['WIDENING_POTENTIAL'].notna()]
        tempDF = tempDF[~tempDF['WIDENING_POTENTIAL'].isin(range(1,5)) | ~tempDF['WIDENING_POTENTIAL_VALUE_TEXT'].str.fullmatch("[XABCDE]")]
        self.df['DRE42'].iloc[tempDF.index.tolist()] = False

    def dre43a(self):
        #CURVES_A ValueNumeric must be in the format Numeric (3;1) and > 0

        print('Running rule DRE43A...')
        self.df['DRE43A'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_A'].notna()]
        tempDF = tempDF[~tempDF['CURVES_A'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43A'].iloc[tempDF.index.tolist()] = False

    def dre43b(self):
        #CURVES_B ValueNumeric must be in the format Numeric (3;1) and > 0

        print('Running rule DRE43B...')
        self.df['DRE43B'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_B'].notna()]
        tempDF = tempDF[~tempDF['CURVES_B'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43B'].iloc[tempDF.index.tolist()] = False

    def dre43c(self):
        #CURVES_C ValueNumeric must be in the format Numeric (3;1) and > 0

        print('Running rule DRE43C...')
        self.df['DRE43C'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_C'].notna()]
        tempDF = tempDF[~tempDF['CURVES_C'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43C'].iloc[tempDF.index.tolist()] = False

    def dre43d(self):
        #CURVES_D ValueNumeric must be in the format Numeric (3;1) and > 0

        print('Running rule DRE43D...')
        self.df['DRE43D'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_D'].notna()]
        tempDF = tempDF[~tempDF['CURVES_D'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43D'].iloc[tempDF.index.tolist()] = False

    def dre43e(self):
        #CURVES_E ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE43E...')
        self.df['DRE43E'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_E'].notna()]
        tempDF = tempDF[~tempDF['CURVES_E'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43E'].iloc[tempDF.index.tolist()] = False

    def dre43f(self):
        #CURVES_F ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE43F...')
        self.df['DRE43F'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CURVES_F'].notna()]
        tempDF = tempDF[~tempDF['CURVES_F'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE43F'].iloc[tempDF.index.tolist()] = False

    def dre44(self):
        #TERRAIN_TYPE ValueNumeric must be an integer in the range (1-3)

        print('Running rule DRE44...')
        self.df['DRE44'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['TERRAIN_TYPE'].notna()]
        tempDF = tempDF[~tempDF['TERRAIN_TYPE'].isin([1,2,3])]
        self.df['DRE44'].iloc[tempDF.index.tolist()] = False      

    def dre45a(self):
        #GRADES_A ValueNumeric must be in the format Numeric (3;1) and > 0

        print('Running rule DRE45A...')
        self.df['DRE45A'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_A'].notna()]
        tempDF = tempDF[~tempDF['GRADES_A'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45A'].iloc[tempDF.index.tolist()] = False


    def dre45b(self):
        #GRADES_B ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE45B...')
        self.df['DRE45B'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_B'].notna()]
        tempDF = tempDF[~tempDF['GRADES_B'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45B'].iloc[tempDF.index.tolist()] = False

    def dre45c(self):
        #GRADES_C ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE45C...')
        self.df['DRE45C'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_C'].notna()]
        tempDF = tempDF[~tempDF['GRADES_C'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45C'].iloc[tempDF.index.tolist()] = False

    def dre45d(self):
        #GRADES_D ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE45D...')
        self.df['DRE45D'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_D'].notna()]
        tempDF = tempDF[~tempDF['GRADES_D'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45D'].iloc[tempDF.index.tolist()] = False


    def dre45e(self):
        #GRADES_E ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE45E...')
        self.df['DRE45E'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_E'].notna()]
        tempDF = tempDF[~tempDF['GRADES_E'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45E'].iloc[tempDF.index.tolist()] = False


    def dre45f(self):
        #GRADES_F ValueNumeric must be in the format Numeric (3;1) and > 0
        print('Running rule DRE45F...')
        self.df['DRE45F'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['GRADES_F'].notna()]
        tempDF = tempDF[~tempDF['GRADES_F'].astype(str).str.fullmatch("[0-9]\.[0-9]{0,3}")]
        self.df['DRE45F'].iloc[tempDF.index.tolist()] = False

    def dre46(self):
        #PCT_PASS_SIGHT ValueNumeric must be an integer in the range (0-100)

        print('Running rule DRE46...')
        self.df['DRE46'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['PCT_PASS_SIGHT'].notna()]
        tempDF = tempDF[~tempDF['PCT_PASS_SIGHT'].isin(range(0,101))]
        self.df['DRE46'].iloc[tempDF.index.tolist()] = False

    def dre47(self):
        #IRI: Where ValueNumeric is not Null; it must be an integer >0; ValueNumeric may be NULL Where ValueText = E; 
        # ValueDate must have a format of MM/DD/YYYY; Where ValueDate <> BeginDate; ValueText must be in (A;B;C;D;E)

        #Site isn't rejecting non integer IRI values, which is roughly half of the data. Removing the check for whether or not 
        #Iri is an integer


        print('Running rule DRE47...')
        self.df['DRE47'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI'].notna()]
        #Checks whether IRI is an integer > 0, not used atm
        # tempDF = tempDF[(tempDF['IRI'] <= 0) | (tempDF['IRI']%np.floor(tempDF['IRI']) != 0)]
        tempDF = tempDF[tempDF['IRI'] <= 0]
        self.df['DRE47'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI_VALUE_TEXT'].notna()]
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['IRI_VALUE_TEXT'].str.fullmatch("E")]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['DRE47'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].notna()]
        tempDF = tempDF[~tempDF['IRI_VALUE_DATE'].astype(str).str.contains(date_regex)]
        self.df['DRE47'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].notna()]
        tempDF['IRI_VALUE_DATE'] = pd.to_datetime(tempDF['IRI_VALUE_DATE'])
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].dt.year != tempDF['BEGIN_DATE'].dt.year]
        tempDF['IRI_VALUE_TEXT'].fillna("", inplace=True)
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['IRI_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE47'].iloc[tempDF.index.tolist()] = False

    def dre48(self):
        #PSR ValueNumeric must be > 0.0 and <= 5.0 and in format Numeric (2;1);
        #  ValueText must be NULL or = A; ValueDate must have a format of MM/DD/YYYY; Where ValueDate <> BeginDate; ValueText must be in (A;B;C;D;E)

        #Skipping rule, we don't report PSR

        print('Running rule DRE48...')
        self.df['DRE48'] = True


    def dre49(self):
        #SURFACE_TYPE ValueNumeric must be an integer in the range (1-11)

        print('Running rule DRE49...')
        self.df['DRE49'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].notna()]
        tempDF = tempDF[~tempDF['SURFACE_TYPE'].isin(range(1,12))]
        self.df['DRE49'].iloc[tempDF.index.tolist()] = False

    def dre50(self):
        #RUTTING: Where ValueNumeric is not Null; it must be an integer >=0; ValueNumeric may be NULL Where ValueText = E; 
        # ValueDate must have a format of MM/DD/YYYY Where ValueDate <> BeginDate ValueText must be in (A;B;C;D;E)

        #Ignoring "must be an integer" as 90% of our data for rutting is not an integer and the HPMS manual states to report to nearest .01 inch

        print('Running rule DRE50...')
        self.df['DRE50'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING'].notna()]
        tempDF = tempDF[(tempDF['RUTTING'] <= 0)]
        self.df['DRE50'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING_VALUE_TEXT'].notna()]
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['RUTTING_VALUE_TEXT'].str.fullmatch("E")]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['DRE50'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].notna()]
        tempDF = tempDF[~tempDF['RUTTING_VALUE_DATE'].astype(str).str.contains("\d{1,2}/\d{1,2}/\d{4}")]
        self.df['DRE50'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].notna()]
        tempDF['RUTTING_VALUE_DATE'] = pd.to_datetime(tempDF['RUTTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].dt.year != tempDF['BEGIN_DATE'].dt.year]
        tempDF['RUTTING_VALUE_TEXT'].fillna("", inplace=True)
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['RUTTING_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE50'].iloc[tempDF.index.tolist()] = False

    def dre51(self):
        #FAULTING: Where ValueNumeric is not Null; it must be an integer >=0; ValueNumeric may be NULL Where ValueText = E; 
        # ValueDate must have a format of MM/DD/YYYY; Where ValueDate <> BeginDate; ValueText must be in (A;B;C;D;E)

        #Ignoring "must be an integer" as 90% of our data for faulting is not an integer and the HPMS manual states to report to nearest .01 inch

        print('Running rule DRE51...')
        self.df['DRE51'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING'].notna()]
        tempDF = tempDF[(tempDF['FAULTING'] <= 0)]
        self.df['DRE51'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING_VALUE_TEXT'].notna()]
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['FAULTING_VALUE_TEXT'].str.fullmatch("E")]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['DRE51'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].notna()]
        tempDF = tempDF[~tempDF['FAULTING_VALUE_DATE'].astype(str).str.contains("\d{1,2}/\d{1,2}/\d{4}")]
        self.df['DRE51'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].notna()]
        tempDF['FAULTING_VALUE_DATE'] = pd.to_datetime(tempDF['FAULTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].dt.year != tempDF['BEGIN_DATE'].dt.year]
        tempDF['FAULTING_VALUE_TEXT'].fillna("", inplace=True)
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['FAULTING_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE51'].iloc[tempDF.index.tolist()] = False

    def dre52(self):
        #CRACKING_PERCENT: Where ValueNumeric is not Null; it must be an integer >=0; ValueNumeric may be NULL Where ValueText = E; 
        # ValueDate must have a format of MM/DD/YYYY; Where ValueDate <> BeginDate; ValueText must be in (A;B;C;D;E)

        #Ignoring "must be an integer" as 90% of our data for cracking_percent is not an integer and the HPMS manual states to report to nearest .01 inch

        print('Running rule DRE52...')
        self.df['DRE52'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].notna()]
        tempDF = tempDF[(tempDF['CRACKING_PERCENT'] <= 0)]
        self.df['DRE52'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_TEXT'].notna()]
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['CRACKING_PERCENT_VALUE_TEXT'].str.fullmatch("E")]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['DRE52'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].notna()]
        tempDF = tempDF[~tempDF['CRACKING_PERCENT_VALUE_DATE'].astype(str).str.contains("\d{1,2}/\d{1,2}/\d{4}")]
        self.df['DRE52'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].notna()]
        tempDF['CRACKING_PERCENT_VALUE_DATE'] = pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE'])
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].dt.year != tempDF['BEGIN_DATE'].dt.year]
        tempDF['CRACKING_PERCENT_VALUE_TEXT'].fillna("", inplace=True)
        if tempDF.shape[0] != 0:
            tempDF = tempDF[~tempDF['CRACKING_PERCENT_VALUE_TEXT'].str.fullmatch("[ABCDE]")]
        self.df['DRE52'].iloc[tempDF.index.tolist()] = False

    def dre54(self):
        #YEAR_LAST_IMPROVEMENT ValueDate must not be NULL and must be <= BeginDate

        print('Running rule DRE54...')
        self.df['DRE54'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['YEAR_LAST_IMPROVEMENT_VALUE_DATE'].isna() | (tempDF['YEAR_LAST_IMPROVEMENT_VALUE_DATE'] > tempDF['BEGIN_DATE'].dt.year)]
        self.df['DRE54'].iloc[tempDF.index.tolist()] = False

    def dre55(self):
        #YEAR_LAST_CONSTRUCTION ValueDate must not be NULL and must be <= BeginDate

        print('Running rule DRE55...')
        self.df['DRE55'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['YEAR_LAST_CONSTRUCTION_VALUE_DATE'].isna() | (tempDF['YEAR_LAST_CONSTRUCTION_VALUE_DATE'] > tempDF['BEGIN_DATE'].dt.year)]
        self.df['DRE55'].iloc[tempDF.index.tolist()] = False

    def dre56(self):
        #LAST_OVERLAY_THICKNESS ValueNumeric must be > 0 

        print('Running rule DRE56...')
        self.df['DRE56'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['LAST_OVERLAY_THICKNESS'].notna()]
        tempDF = tempDF[tempDF['LAST_OVERLAY_THICKNESS'] <= 0]
        self.df['DRE56'].iloc[tempDF.index.tolist()] = False

    def dre57(self):
        #THICKNESS_RIGID ValueNumeric must be > 0

        print('Running rule DRE57...')
        self.df['DRE57'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['THICKNESS_RIGID'].notna()]
        tempDF = tempDF[tempDF['THICKNESS_RIGID'] <= 0]
        self.df['DRE57'].iloc[tempDF.index.tolist()] = False

    def dre58(self):
        #THICKNESS_RIGID ValueNumeric must be > 0

        print('Running rule DRE58...')
        self.df['DRE58'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['THICKNESS_FLEXIBLE'].notna()]
        tempDF = tempDF[tempDF['THICKNESS_FLEXIBLE'] <= 0]
        self.df['DRE58'].iloc[tempDF.index.tolist()] = False

    def dre59(self):
        #BASE_TYPE ValueNumeric must be an integer in the range (1-8)

        print('Running rule DRE59...')
        self.df['DRE59'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['BASE_TYPE'].notna()]
        tempDF = tempDF[~tempDF['BASE_TYPE'].isin(range(1,9))]
        self.df['DRE59'].iloc[tempDF.index.tolist()] = False

    def dre60(self):
        #BASE_THICKNESS ValueNumeric must be an integer > 0

        print('Running rule DRE60...')
        self.df['DRE60'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['BASE_THICKNESS'].notna()]
        tempDF = tempDF[(tempDF['BASE_THICKNESS'] <= 0) | (tempDF['BASE_THICKNESS']%np.floor(tempDF['BASE_THICKNESS']) != 0)]
        self.df['DRE60'].iloc[tempDF.index.tolist()] = False

    def dre62(self):
        #SOIL_TYPE ValueNumeric must be an integer in the range (1;2)

        print('Running rule DRE62...')
        self.df['DRE62'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SOIL_TYPE'].notna()]
        tempDF = tempDF[~tempDF['SOIL_TYPE'].isin([1,2])]
        self.df['DRE62'].iloc[tempDF.index.tolist()] = False

    def dre63(self):
        #COUNTY_ID ValueNumeric must be an integer and valid three digit FIPS code

        print('Running rule DRE63...')
        self.df['DRE63'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[~tempDF['COUNTY_ID'].isin(range(1,110, 2))]
        self.df['DRE63'].iloc[tempDF.index.tolist()] = False

    def dre68(self):
        #MAINTENANCE_OPERATIONS must be an integer in (1;2;3;4;11;12;21;25;26;27;31;32;40;50;60;62;63;64;66;67;68;69;70;72;73;74;80) 

        print('Running rule DRE68...')
        self.df['DRE68'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['MAINTENANCE_OPERATIONS'].notna()]
        tempDF = tempDF[~tempDF['MAINTENANCE_OPERATIONS'].isin([1,2,3,4,11,12,21,25,26,27,31,32,40,50,60,62,63,64,66,67,68,69,70,72,73,74,80])]
        self.df['DRE68'].iloc[tempDF.index.tolist()] = False

    def dre70(self):
        #DIR_THROUGH_LANES ValueNumeric must be an integer > 0 

        print('Running rule DRE70...')
        self.df['DRE70'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'].notna()]
        tempDF = tempDF[(tempDF['DIR_THROUGH_LANES'] <= 0) | (tempDF['DIR_THROUGH_LANES']%np.floor(tempDF['DIR_THROUGH_LANES']) != 0)]
        self.df['DRE70'].iloc[tempDF.index.tolist()] = False

    def dre71(self):
        #TRAVEL_TIME_CODE ValueText must not be NULL

        print('Running rule DRE71...')
        self.df['DRE71'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['TRAVEL_TIME_CODE_VALUE_TEXT'].isna()]
        self.df['DRE71'].iloc[tempDF.index.tolist()] = False

    def dre73(self):
        #IS_RESTRICTED ValueNumeric must = 1 or NULL

        print('Running rule DRE73...')
        self.df['DRE73'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['IS_RESTRICTED'] != 1) & tempDF['IS_RESTRICTED'].notna()]
        self.df['DRE73'].iloc[tempDF.index.tolist()] = False

    def dre74(self):
        #Section Length Must Not Be > 0.11 Miles

        print('Running rule DRE74...')
        self.df['DRE74'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[(tempDF['EMP'] - tempDF['BMP']) > 0.11]
        self.df['DRE74'].iloc[tempDF.index.tolist()] = False

    def dre75(self):
        #Section Length Must Not Be > 0.11 Miles
        #Skipped since we don't report PSR
        print('Running rule DRE75...')
        self.df['DRE75'] = True

    def dre76(self):
        #Section Length Must Not Be > 0.11 Miles

        print('Running rule DRE76...')
        self.df['DRE76'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING'].notna()]
        tempDF = tempDF[(tempDF['EMP'] - tempDF['BMP']) > 0.11]
        self.df['DRE76'].iloc[tempDF.index.tolist()] = False

    def dre77(self):
        #Section Length Must Not Be > 0.11 Miles

        print('Running rule DRE77...')
        self.df['DRE77'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING'].notna()]
        tempDF = tempDF[(tempDF['EMP'] - tempDF['BMP']) > 0.11]
        self.df['DRE77'].iloc[tempDF.index.tolist()] = False

    def dre78(self):
        #Section Length Must Not Be > 0.11 Miles

        print('Running rule DRE78...')
        self.df['DRE78'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].notna()]
        tempDF = tempDF[(tempDF['EMP'] - tempDF['BMP']) > 0.11]
        self.df['DRE78'].iloc[tempDF.index.tolist()] = False


    def run(self):

        self.drd01()
        self.drd64()
        self.drd65()
        self.drd66()
        # self.drd72() Data item not in all submission data
        self.drd202()
        self.drd203()
        self.drd204()
        self.dre02()
        self.dre03()
        self.dre04()
        self.dre05()
        self.dre06()
        self.dre07()
        self.dre10()
        self.dre11()
        self.dre12()
        self.dre13()
        self.dre14()
        self.dre15()
        self.dre21()
        self.dre22()
        self.dre23()
        self.dre24()
        self.dre25()
        self.dre26()
        self.dre27()
        self.dre28()
        self.dre29()
        self.dre30()
        self.dre31()
        self.dre32()
        self.dre33()
        self.dre34()
        self.dre35()
        self.dre36()
        self.dre37()
        self.dre38()
        self.dre39()
        self.dre40()
        self.dre42()
        self.dre43a()
        self.dre43b()
        self.dre43c()
        self.dre43d()
        self.dre43e()
        self.dre43f()
        self.dre44()
        self.dre45a()
        self.dre45b()
        self.dre45c()
        self.dre45d()
        self.dre45e()
        self.dre45f()
        self.dre46()
        self.dre47()
        self.dre48()
        self.dre49()
        self.dre50()
        self.dre51()
        self.dre52()
        self.dre54()
        self.dre55()
        self.dre56()
        self.dre57()
        self.dre58()
        self.dre59()
        self.dre60()
        # self.dre62() Data Item not in all submission data
        self.dre63()
        self.dre68()
        self.dre70()
        # self.dre71() Data Item not in all submission data
        # self.dre73() Data item not in all submission data
        self.dre74()
        self.dre75()
        self.dre76()
        self.dre77()
        self.dre78()

    def create_output(self, template='domain_rules_summary_template.xlsx', outfilename='domain_rules_summary.xlsx'):
        #Reads sheet on template that list all data items associated with each rule and converts to dictionary
        dataItemsDF = pd.read_excel(template, sheet_name="ruleDataItems", usecols='A,B', nrows=86)
        dataItemsDF['Rule'] = dataItemsDF['Rule'].str.replace("-", "")
        dataItemsDF['Data_Items'] = dataItemsDF['Data_Items'].str.split(",")
        ruleDict = dict(zip(dataItemsDF['Rule'], dataItemsDF['Data_Items']))
        fwhaRuleDict = ruleDict.copy()
        for rule in ruleDict.keys():
            fwhaRuleDict[rule+"_FHWA"] = fwhaRuleDict.pop(rule)


        #Reads the rule descripts off of summary sheet and converts to dictionary
        ruleDescDF = pd.read_excel(template, sheet_name="Summary", usecols="A,D")
        ruleDesc = dict(zip(ruleDescDF['Rule'], ruleDescDF['Description']))

        #Create copy of template to write to
        shutil.copy(template, outfilename)

        with pd.ExcelWriter(outfilename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            tempDF = self.df.copy()
            numFailed = []
            numPassed = []
            lenFailed = []
            fwhaFailed = []

            #Get counts for failed/passed/length of failed
            print("Updating summary sheet on",outfilename,"...")
            for rule in ruleDict.keys():
                #Assumes all passes for rules not ran
                try:
                    numFailed.append(tempDF[tempDF[rule]==False].shape[0])
                    numPassed.append(tempDF[tempDF[rule]==True].shape[0])
                    lenFailed.append((tempDF[tempDF[rule]==False]['EMP'] - tempDF[tempDF[rule]==False]['BMP']).sum())
                except KeyError:
                    numFailed.append(0)
                    numPassed.append(self.df.shape[0])
                    lenFailed.append(0)

            for rule in fwhaRuleDict.keys():
                try:
                    fwhaFailed.append(tempDF[tempDF[rule]==False].shape[0])
                except KeyError:
                    fwhaFailed.append(0)

            failedDF = pd.DataFrame(numFailed)
            passedDF = pd.DataFrame(numPassed)
            lengthDF = pd.DataFrame(lenFailed)
            fwhaFailedDF = pd.DataFrame(fwhaFailed)

            #Write counts to Summary sheet
            failedDF.to_excel(writer, sheet_name='Summary', startcol=4, startrow=1, header=False, index=False)
            passedDF.to_excel(writer, sheet_name='Summary', startcol=5, startrow=1, header=False, index=False)
            lengthDF.to_excel(writer, sheet_name='Summary', startcol=6, startrow=1, header=False, index=False)
            fwhaFailedDF.to_excel(writer, sheet_name='Summary', startcol=7, startrow=1, header=False, index=False)

            #Create sheets for each rule containing all failed rows (using only columns that the specific rule references)
            for rule in ruleDict.keys():
                tempDF = self.df.copy()
                #Checks to make sure rule has data items associated with it (will be a list if dataItems exists, otherwise will be float (np.nan))
                if type(ruleDict[rule])==list:
                    #Tries using RULENAME (i.e. SJF01) in dataset which is added if the rule is ran
                    #If rule is not run, no column will exist with the rulename, catches KeyError and prints message.
                    try:
                        if tempDF[tempDF[rule]==False].shape[0] > 0:
                            print("Creating error sheet for rule:",rule)
                            dataItems = ['RouteID', 'BMP', 'EMP']
                            [dataItems.append(x) for x in ruleDict[rule] if x not in dataItems]
                            tempDF = tempDF[tempDF[rule]==False]
                            tempDF = tempDF[dataItems]
                            tempDF.to_excel(writer, sheet_name=rule, startrow=1, index=False)
                            worksheet = writer.sheets[rule]
                            worksheet['A1'] = f'=HYPERLINK("#Summary!A1", "Summary Worksheet")'
                            worksheet['A1'].font = Font(underline='single', color='0000EE')
                            #Autofit columns
                            for i in range(1, worksheet.max_column+1):
                                worksheet.column_dimensions[get_column_letter(i)].width = 20
                            #Add rule description to sheet
                            worksheet['B1'] = ruleDesc[rule]

                        else:
                            print("No failed rows for rule:",rule)

                    except KeyError:
                        print(rule,"not found in DF. Sheet was not created in rules_summary.xlsx")
                else:
                    print("No data items for rule",rule,", Sheet not created.")






df = pd.read_csv("all_submission_data.csv", dtype={'URBAN_CODE':str, 'AADT_VALUE_DATE':str})
c = domain_validations(df)
c.run()
c.create_output()
# c.df.to_csv('test_domain_tyler.csv', index=False) 