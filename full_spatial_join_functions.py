import pandas as pd
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
import numpy as np
import json
import warnings
import copy
import string
import os
import shutil
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")


f_system_dict = {
            1:1,
            11:1,
            4:2,
            12:2,
            2:3,
            14:3,
            6:4,
            16:4,
            7:5,
            17:5,
            8:6,
            18:6,
            9:7,
            19:7
        }

facility_list = [1,2,4,5,6]
facility_list2 = [1,2,4]
f_system_list = [1,2,3,4,5,6,7]
URBAN_CODE_list = ['06139','15481','21745','36190','40753','59275','67672','93592','94726']

column_list = []

# reading in rule information
# rules = pd.read_excel('new_rules_used.xlsx')
# rules['UsedColumns'] = rules.UsedColumns.map(lambda x: json.loads(x.replace("'", '"').replace("URBAN_ID",'URBAN_CODE')))
# rules = rules[(rules.FileName == 'Spatial Join Full')&(rules['Data Item Name'] !='STRUCTURE_TYPE')]
# rules_col_used = rules.set_index('Rule')['UsedColumns'].to_dict()
# rules_description = rules.set_index('Rule')['Message'].to_dict()
# rules_name = rules.set_index('Rule')['Data Item Name'].to_dict()


class full_spatial_functions():
  #if the sjf columsn return false, 
  # it means the data passes that validation, 
  # if the sjf column returns True, the
  # data faios the validation
  # Comments in the code are the actual rule that is being applied in the function  
    
    
    def __init__(self,df):
        self.error_df = pd.DataFrame()
        self.df = df

        try:
            self.df['BEGIN_DATE']
        except KeyError:
            self.df['BEGIN_DATE'] = datetime(2022,1,1)

        self.df.rename(columns = {
            "FsystemVn":"F_SYSTEM",
            "NhsVn":"NHS",
            "NnVn":"NN",
            "UrbanIdVn":"URBAN_CODE",
            "FacilityTypeVn":"FACILITY_TYPE",
            "StructureTypeVn":"STRUCTURE_TYPE",
            "OwnershipVn":"OWNERSHIP",
            "CountyIdVn":"COUNTY_ID",
            "MaintenanceOperationsVn":"MAINTENANCE_OPERATIONS",
            "ThroughLanesVn":"THROUGH_LANES",
            "PeakLanesVn":"PEAK_LANES",
            "CounterPeakLanesVn":"COUNTER_PEAK_LANES",
            "LaneWidthVn":"LANE_WIDTH",
            "MedianTypeVn":"MEDIAN_TYPE",
            "MedianWidthVn":"MEDIAN_WIDTH",
            "ShoulderTypeVn":"SHOULDER_TYPE",
            "ShoulderWidthRVn":"SHOULDER_WIDTH_R",
            "ShoulderWidthLVn":"SHOULDER_WIDTH_L",
            "PeakParkingVn":"PEAK_PARKING",
            "DirThroughLanesVn":"DIR_THROUGH_LANES",
            "TurnLanesRVn":"TURN_LANES_R",
            "TurnLanesLVn":"TURN_LANES_L",
            "SignalTypeVn":"SIGNAL_TYPE",
            "PctGreenTimeVn":"PCT_GREEN_TIME",
            "NumberSignalsVn":"NUMBER_SIGNALS",
            "StopSignsVn":"STOP_SIGNS",
            "AtGradeOtherVn":"AT_GRADE_OTHER",
            "AadtVn":"AADT",
            "AadtVt":"AADT_VALUE_TEXT",
            "AadtVd":"AADT_VALUE_DATE",
            "AadtsingleUnitVn":"AADT_SINGLE_UNIT",
            "AadtsingleUnitVt":"AADT_SINGLE_UNIT_VALUE_TEXT",
            "AadtsingleUnitVd":"AADT_SINGLE_UNIT_VALUE_DATE",
            "AadtcombinationVn":"AADT_COMBINATION",
            "PctdhsingleVn":"PCT_DH_SINGLE_UNIT",
            "PctdhcombinationVn":"PCT_DH_COMBINATION",
            "KfactorVn":"K_FACTOR",
            "DirFactorVn":"DIR_FACTOR",
            "FutureAadtVn":"FUTURE_AADT",
            "AccessControlVn":"ACCESS_CONTROL",
            "SpeedLimitVn":"SPEED_LIMIT",
            "IriVn":"IRI",
            "IriVt":"IRI_VALUE_TEXT",
            "IriVd":"IRI_VALUE_DATE",
            "SurfaceTypeVn":"SURFACE_TYPE",
            "RuttingVn":"RUTTING",
            "RuttingVt":"RUTTING_VALUE_TEXT",
            "RuttingVd":"RUTTING_VALUE_DATE",
            "FaultingVn":"FAULTING",
            "FaultingVt":"FAULTING_VALUE_TEXT",
            "FaultingVd":"FAULTING_VALUE_DATE",
            "CrackingPercentVn":"CRACKING_PERCENT",
            "CrackingPercentVt":"CRACKING_PERCENT_VALUE_TEXT",
            "CrackingPercentVd":"CRACKING_PERCENT_VALUE_DATE",
            "YearLastImprovementVd":"YEAR_LAST_IMPROVEMENT_VALUE_DATE",
            "YearLastConstructionVd":"YEAR_LAST_CONSTRUCTION_VALUE_DATE",
            "LastOverlayThicknessVn":"LAST_OVERLAY_THICKNESS",
            "ThicknessRigidVn":"THICKNESS_RIGID",
            "ThicknessFlexibleVn":"THICKNESS_FLEXIBLE",
            "BaseTypeVn":"BASE_TYPE",
            "BaseThicknessVn":"BASE_THICKNESS",
            "SoilTypeVn":"SOIL_TYPE",
            "WideningPotentialVn":"WIDENING_POTENTIAL",
            "CurvesAVn":"CURVES_A",
            "CurvesBVn":"CURVES_B",
            "CurvesCVn":"CURVES_C",
            "CurvesDVn":"CURVES_D",
            "CurvesEVn":"CURVES_E",
            "CurvesFVn":"CURVES_F",
            "TerrarinTypeVn":"TERRAIN_TYPE",
            "GradesAVn":"GRADES_A",
            "GradesBVn":"GRADES_B",
            "GradesCVn":"GRADES_C",
            "GradesDVn":"GRADES_D",
            "GradesEVn":"GRADES_E",
            "GradesFVn":"GRADES_F",
            "PctPassSightVn":"PCT_PASS_SIGHT",
            "RouteQualifier":"ROUTE_QUALIFIER",
            "RouteSigning":"ROUTE_SIGNING",
            "RouteName":"ROUTE_NAME_VALUE_TEXT",
            "RouteNumber":"ROUTE_NUMBER",
            "SampleId":"HPMS_SAMPLE_NO"
        }, inplace=True)
        
    
    def sjf01(self):
        #F_SYSTEM must exist where FACILITY_TYPE is in (1;2;4;5;6) and Must not be NULL
        print("Running rule SJF01...")
        self.df['SJF01'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4,5,6]) & tempDF['F_SYSTEM'].isna()]
        self.df['SJF01'].iloc[tempDF.index.tolist()] = False

    def sjf02(self):
        #URBAN_CODE must exist and must not be NULL where: 1. FACILITY_TYPE in (1;2;4) AND F_SYSTEM in (1;2;3;4;5) 
        # [OR] 2. FACILITY_TYPE = 6 AND DIR_THROUGH_LANES > 0 and F_SYSTEM = 1 AND (IRI IS NOT NULL OR PSR IS NOT NULL)
        print("Running rule SJF02...")
        self.df['SJF02'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4])]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4,5])]
        tempDF = tempDF[tempDF['URBAN_CODE'].isna()]
        self.df['SJF02'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE']==6]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES']>0]
        tempDF = tempDF[tempDF['F_SYSTEM']==1]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['URBAN_CODE'].isna()]
        self.df['SJF02'].iloc[tempDF.index.tolist()] = False

    def sjf03(self):
        #FACILITY_TYPE must exist where F_SYSTEM in (1;2;3;4;5;6;7)  and must not be NULL
        print("Running rule SJF03...")
        self.df['SJF03'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin(range(1,8)) & tempDF['FACILITY_TYPE'].isna()]
        self.df['SJF03'].iloc[tempDF.index.tolist()] = False

    def sjf04(self):
        #No validation
        print("Running rule SJF04...")
        self.df['SJF04'] = True

    def sjf05(self):
        #ACCESS_CONTROL must exist where (F_SYSTEM in (1;2;3) or Sample or NHS) AND FACILITY_TYPE IN (1;2) and must not be NULL
        print("Running rule SJF05...")
        self.df['SJF05'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | tempDF['HPMS_SAMPLE_NO'].notna() | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['ACCESS_CONTROL'].isna()]
        self.df['SJF05'].iloc[tempDF.index.tolist()] = False


    def sjf06(self):
        #OWNERSHIP must exist where (F_SYSTEM in (1;2;3;4;5;6;7) and FACILITY_TYPE (1;2;5;6) and must not be NULL
        print("Running rule SJF06...")
        self.df['SJF06'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin(range(1,8))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,5,6])]
        tempDF = tempDF[tempDF['OWNERSHIP'].isna()]
        self.df['SJF06'].iloc[tempDF.index.tolist()] = False

    def sjf07(self):
        #THROUGH_LANES must exist where FACILITY_TYPE in (1;2;4) 
        # AND (F_SYSTEM in (1;2;3;4;5) or (F_SYSTEM = 6 and URBAN_CODE <99999) or NHS ValueNumeric <> NULL) and must not be NULL
        #Q or (R AND S) or T === (Q or R or T) AND (Q or S or T)
        print("Running rule SJF07...")
        self.df['SJF07'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4])]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin(range(1,6)) | (tempDF['F_SYSTEM'] == 6) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin(range(1,6)) | (tempDF['URBAN_CODE'].astype(float) < 99999) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['THROUGH_LANES'].isna()]
        self.df['SJF07'].iloc[tempDF.index.tolist()] = False

    def sjf08(self):
        #MANAGED_LANES_TYPE must exist where MANAGED_LANES is not Null
        print("Running rule SJF08...")
        self.df['SJF08'] = True
        # tempDF = self.df.copy()
        # tempDF[tempDF['MANAGED_LANES'].notna() & tempDF['MANAGED_LANES_TYPE'].isna()]
        # self.df['SJF08'].iloc[tempDF.index.tolist()] = False

    def sjf09(self):
        #MANAGED_LANES must exist where MANAGED_LANES_TYPE is not Null
        print("Running rule SJF09...")
        self.df['SJF09'] = True
        # tempDF = self.df.copy()
        # tempDF[tempDF['MANAGED_LANES_TYPE'].notna() & tempDF['MANAGED_LANES'].isna()]
        # self.df['SJF09'].iloc[tempDF.index.tolist()] = False

    def sjf10(self):
        #PEAK_LANES must exist on Samples
        print("Running rule SJF10...")
        self.df['SJF10'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna() & tempDF['PEAK_LANES'].isna()]
        self.df['SJF10'].iloc[tempDF.index.tolist()] = False    

    def sjf11(self):
        #COUNTER_PEAK_LANES must exist on Samples where FACILITY_TYPE = 2 AND (URBAN_CODE < 99999 OR THROUGH_LANES >=4)
        print("Running rule SJF11...")
        self.df['SJF11'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE']==2]
        tempDF = tempDF[(tempDF['URBAN_CODE'].astype(float)<99999) | (tempDF['THROUGH_LANES'] >= 4)]
        tempDF = tempDF[tempDF['COUNTER_PEAK_LANES'].isna()]
        self.df['SJF11'].iloc[tempDF.index.tolist()] = False

    def sjf12(self):
        #TURN_LANES_R must exist on Samples where URBAN_CODE  < 99999 and ACCESS_CONTROL >1
        print("Running rule SJF12...")
        self.df['SJF12'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(tempDF['URBAN_CODE'].astype(float) < 99999) & (tempDF['ACCESS_CONTROL'] > 1)]
        tempDF = tempDF[tempDF['TURN_LANES_R'].isna()]
        self.df['SJF12'].iloc[tempDF.index.tolist()] = False

    def sjf13(self):
        #TURN_LANES_L must exist on Samples where URBAN_CODE  < 99999 and ACCESS_CONTROL >1
        print("Running rule SJF13...")
        self.df['SJF13'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(tempDF['URBAN_CODE'].astype(float) < 99999) & (tempDF['ACCESS_CONTROL'] > 1)]
        tempDF = tempDF[tempDF['TURN_LANES_L'].isna()]
        self.df['SJF13'].iloc[tempDF.index.tolist()] = False

    def sjf14(self):
        #SPEED_LIMIT must exist on Samples and  the NHS
        print("Running rule SJF14...")
        self.df['SJF14'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['SPEED_LIMIT'].isna()]
        self.df['SJF14'].iloc[tempDF.index.tolist()] = False

    def sjf15(self):
        #No validation
        print("Running rule SJF15...")
        self.df['SJF15'] = True

    def sjf16(self):
        #ROUTE_NUMBER ValueNumeric Must Exist where (F_SYSTEM in (1;2;3;4) or NHS ValueNumeric <> NULL ) and FACILITY_TYPE (1;2) and ROUTE_SIGNING in (2;3;4;5;6;7;8;9)  
        # OR F_SYSTEM=1 AND FACILITY_TYPE=6 AND DIR_THROUGH_LANES > 0 AND (IRI IS NOT NULL OR PSR IS NOT NULL)
        print("Running rule SJF16...")
        self.df['SJF16'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4]) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['ROUTE_SIGNING'].isin(range(2,10))]
        tempDF = tempDF[tempDF['ROUTE_NUMBER'].isna()]
        self.df['SJF16'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1 ]
        tempDF = tempDF[tempDF['FACILITY_TYPE'] == 6]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > 0]
        # tempDF = tempDF[tempDF['IRI'].notna() | tempDF['PSR'].notna()]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['ROUTE_NUMBER'].isna()]
        self.df['SJF16'].iloc[tempDF.index.tolist()] = False

    def sjf17(self):
        #ROUTE_SIGNING must exist where (F_SYSTEM in (1;2;3;4) or NHS) and FACILITY_TYPE (1;2)
        print("Running rule SJF17...")
        self.df['SJF17'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4]) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['ROUTE_SIGNING'].isna()]
        self.df['SJF17'].iloc[tempDF.index.tolist()] = False   

    def sjf18(self):
        #ROUTE_QUALIFIER must exist where (F_SYSTEM in (1;2;3;4) or NHS) and FACILITY_TYPE (1;2)
        print("Running rule SJF18...")
        self.df['SJF18'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4]) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['ROUTE_QUALIFIER'].isna()]
        self.df['SJF18'].iloc[tempDF.index.tolist()] = False   

    def sjf19(self):
        #ROUTE_NAME must exist where (F_SYSTEM in (1;2;3;4) or NHS) and FACILITY_TYPE (1;2)
        print("Running rule SJF19...")
        self.df['SJF19'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4]) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['ROUTE_NAME_VALUE_TEXT'].isna()]
        self.df['SJF19'].iloc[tempDF.index.tolist()] = False

    def sjf20(self):    
        #AADT must exist WHERE: (FACILITY_TYPE in (1;2;4) AND (F_SYSTEM in (1;2;3;4;5)) OR (F_SYSTEM = 6 and URBAN_CODE  <99999) OR NHS ValueNumeric <> NULL
        print("Running rule SJF20...")
        self.df['SJF20'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4])]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin(range(1,6))]
        tempDF = tempDF[tempDF['AADT'].isna()]
        self.df['SJF20'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4])]
        tempDF = tempDF[(tempDF['F_SYSTEM']==6)]
        tempDF = tempDF[(tempDF['URBAN_CODE'].astype(float) < 99999)]
        tempDF = tempDF[tempDF['AADT'].isna()]
        self.df['SJF20'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,4])]
        tempDF = tempDF[tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['AADT'].isna()]
        self.df['SJF20'].iloc[tempDF.index.tolist()] = False



    def sjf21(self):
        #AADT_SINGLE_UNIT must exist WHERE ((F_SYSTEM in (1) or NHS ValueNumeric <> NULL) and FACILITY_TYPE (1;2)) and on Samples
        print("Running rule SJF21...")
        self.df['SJF21'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['AADT_SINGLE_UNIT'].isna()]
        self.df['SJF21'].iloc[tempDF.index.tolist()] = False
        
        tempDF = self.df.copy() 
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['AADT_SINGLE_UNIT'].isna()]
        self.df['SJF21'].iloc[tempDF.index.tolist()] = False

    def sjf22(self):
        #PCT_DH_SINGLE_UNIT must exist on Samples
        print("Running rule SJF22...")
        self.df['SJF22'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['PCT_DH_SINGLE_UNIT'].isna()]
        self.df['SJF22'].iloc[tempDF.index.tolist()] = False

    def sjf23(self):
        #AADT_COMBINATION must exist WHERE ((F_SYSTEM in (1) or NHS ValueNumeric <> NULL) and FACILITY_TYPE (1;2)) and on Samples
        print("Running rule SJF23...")
        self.df['SJF23'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['AADT_COMBINATION'].isna()]
        self.df['SJF23'].iloc[tempDF.index.tolist()] = False
        
        tempDF = self.df.copy() 
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | tempDF['NHS'].notna()]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['AADT_COMBINATION'].isna()]
        self.df['SJF23'].iloc[tempDF.index.tolist()] = False

    def sjf24(self):
        #PCT_DH_COMBINATION must exist on Samples
        print("Running rule SJF24...")
        self.df['SJF24'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna() & tempDF['PCT_DH_COMBINATION'].isna()]
        self.df['SJF24'].iloc[tempDF.index.tolist()] = False

    def sjf25(self):
        #K_FACTOR must exist on Samples
        print("Running rule SJF25...")
        self.df['SJF25'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna() & tempDF['K_FACTOR'].isna()]
        self.df['SJF25'].iloc[tempDF.index.tolist()] = False

    def sjf26(self):
        #DIR_FACTOR must exist on Samples
        print("Running rule SJF26...")
        self.df['SJF26'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna() & tempDF['DIR_FACTOR'].isna()]
        self.df['SJF26'].iloc[tempDF.index.tolist()] = False

    def sjf27(self):
        #FUTURE_AADT must exist on Samples
        print("Running rule SJF27...")
        self.df['SJF27'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna() & tempDF['FUTURE_AADT'].isna()]
        self.df['SJF27'].iloc[tempDF.index.tolist()] = False

    def sjf28(self):
        #SIGNAL_TYPE must exist on Samples WHERE (URBAN_CODE <> 99999 AND NUMBER_SIGNALS >=1)
        print("Running rule SJF28...")
        self.df['SJF28'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float) != 99999]
        tempDF = tempDF[tempDF['NUMBER_SIGNALS'] >= 1]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['SIGNAL_TYPE'].isna()]
        self.df['SJF28'].iloc[tempDF.index.tolist()] = False 

    def sjf29(self):
        #PCT_GREEN_TIME must exist on Samples WHERE (NUMBER_SIGNALS >=1 AND URBAN_CODE <99999)
        print("Running rule SJF29...")
        self.df['SJF29'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float) < 99999]
        tempDF = tempDF[tempDF['NUMBER_SIGNALS'] >= 1]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['PCT_GREEN_TIME'].isna()]
        self.df['SJF29'].iloc[tempDF.index.tolist()] = False

    def sjf30(self):
        #NUMBER_SIGNALS must exist on Samples WHERE SIGNAL_TYPE IN (1;2;3;4)
        print("Running rule SJF30...")
        self.df['SJF30'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SIGNAL_TYPE'].isin([1,2,3,4])]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['NUMBER_SIGNALS'].isna()]
        self.df['SJF30'].iloc[tempDF.index.tolist()] = False  

    def sjf31(self):
        #STOP_SIGNS (the number of stop sign controlled intersections) must exist on Samples
        print("Running rule SJF31...")
        self.df['SJF31'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['STOP_SIGNS'].isna()]
        self.df['SJF31'].iloc[tempDF.index.tolist()] = False

    def sjf32(self):
        #AT_GRADE_OTHER (the number of intersections; type 'other') must exist on Samples
        print("Running rule SJF32...")
        self.df['SJF32'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['AT_GRADE_OTHER'].isna()]
        self.df['SJF32'].iloc[tempDF.index.tolist()] = False 

    def sjf33(self):
        #LANE_WIDTH must exist on Samples
        print("Running rule SJF33...")
        self.df['SJF33'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['LANE_WIDTH'].isna()]
        self.df['SJF33'].iloc[tempDF.index.tolist()] = False

    def sjf34(self):
        #MEDIAN_TYPE must exist on Samples
        print("Running rule SJF34...")
        self.df['SJF34'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].isna()]
        self.df['SJF34'].iloc[tempDF.index.tolist()] = False

    def sjf35(self):
        #MEDIAN_WIDTH must exist on Samples where MEDIAN_TYPE in (2;3;4;5;6;7)
        print("Running rule SJF35...")
        self.df['SJF35'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].isin(range(2,8))]
        tempDF = tempDF[tempDF['MEDIAN_WIDTH'].isna()]
        self.df['SJF35'].iloc[tempDF.index.tolist()] = False

    def sjf36(self):
        #SHOULDER_TYPE must exist on Samples
        print("Running rule SJF36...")
        self.df['SJF36'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['SHOULDER_TYPE'].isna()]
        self.df['SJF36'].iloc[tempDF.index.tolist()] = False

    def sjf37(self):
        #SHOULDER_WIDTH_R must exist on Samples where SHOULDER_TYPE in (2;3;4;5;6)
        print("Running rule SJF37...")
        self.df['SJF37'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['SHOULDER_TYPE'].isin(range(2,7))]
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_R'].isna()]
        self.df['SJF37'].iloc[tempDF.index.tolist()] = False

    def sjf38(self):
        #SHOULDER_WIDTH_L must exist on Samples where (SHOULDER_TYPE in (2;3;4;5;6) and MEDIAN_TYPE in (2;3;4;5;6;7))
        print("Running rule SJF38...")
        self.df['SJF38'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['SHOULDER_TYPE'].isin(range(2,7))]
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].isin(range(2,8))]
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_L'].isna()]
        self.df['SJF38'].iloc[tempDF.index.tolist()] = False

    def sjf39(self):
        #PEAK_PARKING must exist on Samples where URBAN_CODE < 99999
        print("Running rule SJF39...")
        self.df['SJF39'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float) < 99999]
        tempDF = tempDF[tempDF['PEAK_PARKING'].isna()]
        self.df['SJF39'].iloc[tempDF.index.tolist()] = False

    def sjf40(self):
        #WIDENING_POTENTIAL must exist on Samples
        print("Running rule SJF40...")
        self.df['SJF40'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['WIDENING_POTENTIAL'].isna()]
        self.df['SJF40'].iloc[tempDF.index.tolist()] = False

    def sjf41(self):
        #CURVES BP/EP on F_SYSTEM in (1;2;3) or F_SYSTEM = 4 and URBAN_CODE = 99999 and SURFACE_TYPE > 1 Must Align with Sample BP/EP
        print("Running rule SJF41...")
        self.df['SJF41'] = True


    def sjf42(self):
        #At least one CURVES_A-F must be coded for each Sample WHERE (F_SYSTEM in (1;2;3) or F_SYSTEM = 4 and URBAN_CODE = 99999) and SURFACE_TYPE > 1.
        print("Running rule SJF42...")
        self.df['SJF42'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | (tempDF['F_SYSTEM'] == 4)]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | (tempDF['URBAN_CODE'].astype(float) == 99999)]
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['CURVES_A'].isna() & tempDF['CURVES_B'].isna() & tempDF['CURVES_C'].isna() & tempDF['CURVES_D'].isna() & tempDF['CURVES_E'].isna() & tempDF['CURVES_F'].isna()] 
        self.df['SJF42'].iloc[tempDF.index.tolist()] = False

    def sjf43(self):
        #Sum Length (CURVES_A + CURVES_B + CURVES_C + CURVES_D + CURVES_E + CURVES_E) Must Equal to the Sample Length on 
        # (Sample and (F_SYSTEM (1;2;3) or (F_SYSTEM = 4 and URBAN_CODE = 99999)
        print("Running rule SJF43...")
        self.df['SJF43'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | ((tempDF['F_SYSTEM'] == 4) & tempDF['URBAN_CODE'] == 99999)]

        for i in ['A', 'B', 'C', 'D', 'E', 'F']:
            tempDF["CURVES_"+i].fillna(0, inplace=True)

        tempDF['sumCurves'] = tempDF['CURVES_A'] + tempDF['CURVES_B'] + tempDF['CURVES_C'] + tempDF['CURVES_D'] + tempDF['CURVES_E'] + tempDF['CURVES_F']
        
        for sampleID in tempDF['HPMS_SAMPLE_NO'].unique():
            temp2 = tempDF.copy()
            temp2 = temp2[temp2['HPMS_SAMPLE_NO'] == sampleID]
            sectionLen = temp2['EMP'].max() - temp2['BMP'].min()
            temp2 = temp2[round(temp2['sumCurves'],3) != round(sectionLen,3)]
            self.df['SJF43'].iloc[temp2.index.tolist()] = False
            
            

    def sjf44(self):
        #TERRAIN_TYPE must exist on Samples WHERE (URBAN_CODE = 99999 AND F_SYSTEM in (1;2;3;4;5))
        print("Running rule SJF44...")
        self.df['SJF44'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float)==99999]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3,4,5])]
        tempDF = tempDF[tempDF['TERRAIN_TYPE'].isna()]
        self.df['SJF44'].iloc[tempDF.index.tolist()] = False

    def sjf45(self):
        #GRADES BP/EP on F_SYSTEM in (1;2;3) or F_SYSTEM = 4 and URBAN_CODE = 99999 and SURFACE_TYPE > 1 Must Align with Sample BP/EP
        #Rule ignored: If Grades exist according to rule SJF46, then this will always be true (with how all submission data is built)
        print("Running rule SJF45...")
        self.df['SJF45'] = True

    def sjf46(self):
        #At least one GRADES_A-F must be coded for each Sample WHERE F_SYSTEM in (1;2;3) or F_SYSTEM = 4 and URBAN_CODE = 99999 and SURFACE_TYPE > 1.
        print("Running rule SJF46...")
        self.df['SJF46'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | (tempDF['F_SYSTEM'] == 4)]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | (tempDF['URBAN_CODE'].astype(float) == 99999)]
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['GRADES_A'].isna() & tempDF['GRADES_B'].isna() & tempDF['GRADES_C'].isna() & tempDF['GRADES_D'].isna() & tempDF['GRADES_E'].isna() & tempDF['GRADES_F'].isna()]
        self.df['SJF46'].iloc[tempDF.index.tolist()] = False

    def sjf47(self):
        #Sum Length (GRADES_A + GRADES_B + GRADES_C + GRADES_D + GRADES_E + GRADES_E) Must Equal to the Sample Length on (Sample and (F_SYSTEM (1;2;3) or (F_SYSTEM = 4 and URBAN_CODE = 99999)))
        print("Running rule SJF47...")
        self.df['SJF47'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | ((tempDF['F_SYSTEM'] == 4) & tempDF['URBAN_CODE'] == 99999)]

        for i in ['A', 'B', 'C', 'D', 'E', 'F']:
            tempDF["GRADES_"+i].fillna(0, inplace=True)

        tempDF['sumGrades'] = tempDF['GRADES_A'] + tempDF['GRADES_B'] + tempDF['GRADES_C'] + tempDF['GRADES_D'] + tempDF['GRADES_E'] + tempDF['GRADES_F']
        
        for sampleID in tempDF['HPMS_SAMPLE_NO'].unique():
            temp2 = tempDF.copy()
            temp2 = temp2[temp2['HPMS_SAMPLE_NO'] == sampleID]
            sectionLen = temp2['EMP'].max() - temp2['BMP'].min()
            temp2 = temp2[round(temp2['sumGrades'],3) != round(sectionLen,3)]
            self.df['SJF47'].iloc[temp2.index.tolist()] = False



    def sjf48(self):
        #PCT_PASS_SIGHT must exist on Samples WHERE: (URBAN_CODE = 99999 and THROUGH_LANES =2 and MEDIAN_TYPE in (1;2) and SURFACE_TYPE > 1)
        print("Running rule SJF48...")
        self.df['SJF48'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float)==99999]
        tempDF = tempDF[tempDF['THROUGH_LANES'] == 2]
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['PCT_PASS_SIGHT'].isna()]
        self.df['SJF48'].iloc[tempDF.index.tolist()] = False

    def sjf49(self):
        #IRI ValueNumeric Must Exist Where SURFACE_TYPE >1 AND (FACILITY_TYPE IN (1;2) AND (PSR ValueText <> 'A' AND (F_SYSTEM in (1;2;3) OR NHS ValueNumeric <>1) 
        # OR Sample sections WHERE (F_SYSTEM = 4 and URBAN_CODE = 99999)OR DIR_THROUGH_LANES >0
        print("Running rule SJF49...")
        self.df['SJF49'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['F_SYSTEM'].isin([1,2,3]) | (tempDF['NHS'] != 1)]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['SJF49'].iloc[tempDF.index.tolist()] = False  

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(tempDF['F_SYSTEM'] == 4) | (tempDF['DIR_THROUGH_LANES'] > 0)]
        tempDF = tempDF[(tempDF['URBAN_CODE'].astype(float) == 99999) | (tempDF['DIR_THROUGH_LANES'] > 0)]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['SJF49'].iloc[tempDF.index.tolist()] = False  

    def sjf50(self):
        #PSR ValueNumeric Must Exist Where IRI ValueNumeric IS NULL AND FACILITY_TYPE IN (1;2) AND SURFACE_TYPE >1 
        # AND(Sample exists AND (F_SYSTEM in (4;6) AND URBAN_CODE <99999 OR F_SYSTEM = 5) 
        # OR (F_SYSTEM = 1 or NHS ValueNumeric <>NULL) AND PSR ValueText = ‘A’)
        #PSR NOT CHECKED
        print("Running rule SJF50...")
        self.df['SJF50'] = True


    def sjf51(self):
        #SURACE_TYPE|"SURFACE_TYPE ValueNumeric Must Exist Where FACILITY_TYPE in (1;2) AND (F_SYSTEM = 1 OR NHS ValueNumeric <> NULL OR Sample exists) OR DIR_THROUGH_LANES >0 AND (IRI IS NOT NULL OR PSR IS NOT NULL) "
        print("Running rule SJF51...")
        self.df['SJF51'] = True
        tempDF = self.df.copy(deep = True)
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | (tempDF['NHS'].notna()) | (tempDF['HPMS_SAMPLE_NO'].notna())]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isna()]
        self.df['SJF51'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > 0]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isna()]
        self.df['SJF51'].iloc[tempDF.index.tolist()] = False


    
    def sjf52(self):
        #RUTTING|"RUTTING ValueNumeric Must Exist Where SURFACE_TYPE in (2;6;7;8) AND (FACILITY_TYPE in (1;2) AND (F_SYSTEM = 1 OR NHS OR Sample) OR DIR_THROUGH_LANES >0 AND (IRI IS NOT NULL OR PSR IS NOT NULL))
        print("Running rule SJF52...")
        self.df['SJF52'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | (tempDF['NHS'].notna()) | (tempDF['HPMS_SAMPLE_NO'].notna())]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['SJF52'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > 0]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['SJF52'].iloc[tempDF.index.tolist()] = False
    
    def sjf53(self):
        # Faulting ValueNumeric Must Exist Where SURFACE_TYPE in (3;4;9;10) AND (FACILITY_TYPE in (1;2)  AND  (F_SYSTEM = 1 OR NHS OR Sample) OR  DIR_THROUGH_LANES >0 AND (IRI IS NOT NULL OR PSR IS NOT NULL))
        print("Running rule SJF53...")
        self.df['SJF53'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,9,10])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | (tempDF['NHS'].notna()) | (tempDF['HPMS_SAMPLE_NO'].notna())]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['SJF53'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,9,10])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > 0]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['SJF53'].iloc[tempDF.index.tolist()] = False

    
    def sjf54(self):
        #Data Item:CRACKING_PERCENT
        #SURFACE_TYPE in (2;3;4;5;6;7;8;9;10) AND (FACILITY_TYPE in (1;2) AND (F_SYSTEM = 1 OR  NHS  OR Sample) OR (DIR_THROUGH_LANES >0 AND (IRI IS NOT NULL OR PSR IS NOT NULL)))
        print("Running rule SJF54...")
        self.df['SJF54'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,3,4,5,6,7,8,9,10])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[(tempDF['F_SYSTEM']==1) | (tempDF['NHS'].notna()) | (tempDF['HPMS_SAMPLE_NO'].notna())]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['SJF54'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,3,4,5,6,7,8,9,10])]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > 0]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['SJF54'].iloc[tempDF.index.tolist()] = False


    def sjf55(self):
        # YEAR_LAST_IMPROVEMENT must exist on Samples where SURFACE_TYPE is in the range (2;3;4;5;6;7;8;9;10) OR where  (YEAR_LAST_CONSTRUCTION < BeginDate Year - 20)
        print("Running rule SJF55...")
        self.df['SJF55'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,3,4,5,6,7,8,9,10]) | (tempDF['YEAR_LAST_CONSTRUCTION_VALUE_DATE'] < (tempDF['BEGIN_DATE'].dt.year - 20))]
        tempDF = tempDF[tempDF['YEAR_LAST_IMPROVEMENT_VALUE_DATE'].isna()]
        self.df['SJF55'].iloc[tempDF.index.tolist()] = False

    def sjf56(self):
        # YEAR_LAST_CONSTRUCTION	YEAR_LAST_CONSTRUCTION must exist on Samples where SURFACE_TYPE is in the range (2;3;4;5;6;7;8;9;10)
        print("Running rule SJF56...")
        self.df['SJF56'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,3,4,5,6,7,8,9,10])]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['YEAR_LAST_CONSTRUCTION_VALUE_DATE'].isna()]
        self.df['SJF56'].iloc[tempDF.index.tolist()] = False

    def sjf57(self):
        # LAST_OVERLAY_THICKNESS	Sample and YEAR_LAST_IMPROVEMENT exists	
        print("Running rule SJF57...")
        self.df['SJF57'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['YEAR_LAST_IMPROVEMENT_VALUE_DATE'].notna()]
        tempDF = tempDF[tempDF['LAST_OVERLAY_THICKNESS'].isna()]
        self.df['SJF57'].iloc[tempDF.index.tolist()] = False

    def sjf58(self):
        # THICKNESS_RIGID	SURFACE_TYPE (3;4;5;7;8;9;10) and Sample
        print("Running rule SJF58...")
        self.df['SJF58'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,5,7,8,9,10])]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['THICKNESS_RIGID'].isna()]
        self.df['SJF58'].iloc[tempDF.index.tolist()] = False
        
    def sjf59(self):
        # THICKNESS_FLEXIBLE SURFACE_TYPE (2;6;7;8) and Sample
        print("Running rule SJF59...")
        self.df['SJF59'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['THICKNESS_FLEXIBLE'].isna()]
        self.df['SJF59'].iloc[tempDF.index.tolist()] = False

    def sjf60(self):
        # BASE_TYPE	Sample and SURFACE_TYPE >1
        print("Running rule SJF60...")
        self.df['SJF60'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'] > 1]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['BASE_TYPE'].isna()]
        self.df['SJF60'].iloc[tempDF.index.tolist()] = False

    def sjf61(self):
        # BASE_THICKNESS	Where BASE_TYPE >1; SURFACE_TYPE >1  and Sample
        print("Running rule SJF61...")
        self.df['SJF61'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[tempDF['BASE_TYPE']>1]
        tempDF = tempDF[tempDF['SURFACE_TYPE']>1]
        tempDF = tempDF[tempDF['BASE_THICKNESS'].isna()]
        self.df['SJF61'].iloc[tempDF.index.tolist()] = False

    def sjf62(self):
        # SOIL TYPE DO NOT REPORT
        self.df['SJF62'] = True
    
    def sjf63(self):
        # COUNTY_ID	FACILITY_TYPE in (1;2) AND (F_SYSTEM in (1;2;3;4;5) or (F_SYSTEM = 6 and URBAN_ID <99999) or NHS
        print("Running rule SJF63...")
        self.df['SJF63'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[(tempDF['F_SYSTEM'].isin([1,2,3,4,5]))|((tempDF['F_SYSTEM']==6)&(tempDF['URBAN_CODE'].astype(float)<99999))|(tempDF['NHS'].notna())]
        tempDF = tempDF[tempDF['COUNTY_ID'].isna()]
        self.df['SJF63'].iloc[tempDF.index.tolist()] = False

    def sjf64(self):
        # NHS	(F_SYSTEM = 1 AND (FACILITY_TYPE  in 1;2;6))
        print("Running rule SJF64...")
        self.df['SJF64'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM']==1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['NHS'].isna()]
        self.df['SJF64'].iloc[tempDF.index.tolist()] = False

    def sjf65(self):
        # STRAHNET DO NOT VALIDATE
        print("Running rule SJF65...")
        self.df['SJF65'] = True

    def sjf66(self):
        # NN DO NOT VALIDATE
        print("Running rule SJF66...")
        self.df['SJF66'] = True

    def sjf67(self):
        # MAINTENANCE_OPERATIONS DO NOT VALIDATE
        print("Running rule SJF67...")
        self.df['SJF67'] = True

    def sjf68(self):
        # DIR_THROUGH_LANES	F_SYSTEM =1 AND (FACILITY_TYPE = 6) AND (IRI OR PSR >0)
        print("Running rule SJF68...")
        self.df['SJF68'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM']==1]
        tempDF = tempDF[tempDF['FACILITY_TYPE']==6]
        tempDF = tempDF[tempDF['IRI'].notna()]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'].isna()]
        self.df['SJF68'].iloc[tempDF.index.tolist()] = False

    def sjf69(self):
        # THROUGH_LANES	THROUGH_LANES must be >1 WHERE FACILITY_TYPE = 2
        print("Running rule SJF69...")
        self.df['SJF69'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE'] == 2]
        tempDF = tempDF[tempDF['THROUGH_LANES'] <= 1 | tempDF['THROUGH_LANES'].isna()]
        self.df['SJF69'].iloc[tempDF.index.tolist()] = False

    def sjf70(self):
        # THROUGH_LANES	The sum of COUNTER_PEAK_LANES + PEAK_LANES must be >= THROUGH_LANES on Samples
        print("Running rule SJF70...")
        self.df['SJF70'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF['COUNTER_PEAK_LANES'].fillna(0, inplace=True)
        tempDF['PEAK_LANES'].fillna(0, inplace=True)
        tempDF['sumCPL_PL'] = tempDF['COUNTER_PEAK_LANES'] + tempDF['PEAK_LANES']
        tempDF = tempDF[tempDF['THROUGH_LANES'].notna()]
        tempDF = tempDF[tempDF['sumCPL_PL'] < tempDF['THROUGH_LANES']]
        self.df['SJF70'].iloc[tempDF.index.tolist()] = False

    def sjf71(self):
        # COUNTER_PEAK_LANES	COUNTER_PEAK_LANES must be NULL if FACILITY_TYPE is 1
        print("Running rule SJF71...")
        self.df['SJF71'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FACILITY_TYPE']==1]
        tempDF = tempDF[tempDF['COUNTER_PEAK_LANES'].notna()]
        self.df['SJF71'].iloc[tempDF.index.tolist()]=False

    def sjf72(self):
        #SPEED_LIMIT ValueNumeric must be divisible by 5; and < 90 OR = 999
        print("Running rule SJF72...")
        self.df['SJF72'] = True
        tempDF = self.df.copy()
        
        tempDF = tempDF[tempDF['SPEED_LIMIT'].notna()]
        tempDF = tempDF[((tempDF['SPEED_LIMIT'].astype(int)%5) != 0) | ((tempDF['SPEED_LIMIT'] >= 90) & (tempDF['SPEED_LIMIT'] != 999))]
        self.df['SJF72'].iloc[tempDF.index.tolist()] = False

    def sjf73(self):
        # SIGNAL_TYPE	Where F_SYSTEM = 1 and URBAN_ID <> 99999; SIGNAL_TYPE must = 5
        print('Running rule SJF73...')
        self.df['SJF73'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM']==1]
        tempDF = tempDF[tempDF['URBAN_CODE'].astype(float)!=99999]
        tempDF = tempDF[tempDF['SIGNAL_TYPE']!=5]
        self.df['SJF73'].iloc[tempDF.index.tolist()] = False

    def sjf74(self):
        # LANE_WIDTH	LANE_WIDTH ValueNumeric should be > 5 and <19
        print('Running rule SJF74...')
        self.df['SJF74'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['LANE_WIDTH']<=5)|(tempDF['LANE_WIDTH']>=19)]
        tempDF = tempDF[tempDF['LANE_WIDTH'].notna()]
        self.df['SJF74'].iloc[tempDF.index.tolist()] = False

    def sjf75(self):
        # MEDIAN_TYPE	Where MEDIAN_TYPE is in the range (2;3;4;5;6) THEN MEDIAN_WIDTH MUST BE > 0
        print('Running rule SJF75...')
        self.df['SJF75'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['MEDIAN_TYPE'].isin([2,3,4,5,6])]
        tempDF = tempDF[tempDF['MEDIAN_WIDTH']<=0]
        self.df['SJF75'].iloc[tempDF.index.tolist()] = False

    def sjf76(self):
        # MEDIAN_WIDTH	MEDIAN_WIDTH should be NULL if (FACILITY_TYPE ValueNumeric is = 1 or=  4; OR WHERE MEDIAN_TYPE ValueNumeric = 1
        print('Running rule SJF76...')
        self.df['SJF76'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['FACILITY_TYPE'].isin([1,4]))| (tempDF['MEDIAN_TYPE']==1)]
        tempDF = tempDF[tempDF['MEDIAN_WIDTH'].notna()]
        self.df['SJF76'].iloc[tempDF.index.tolist()]= False

    def sjf77(self):
        # SHOULDER_WIDTH_L	SHOULDER_WIDTH_L should be < Median_Width
        print('Running rule SJF77...')
        self.df['SJF77'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SHOULDER_WIDTH_L']>=tempDF['MEDIAN_WIDTH']]
        self.df['SJF77'].iloc[tempDF.index.tolist()] = False
        
    def sjf78(self):
        #IRI
        #ValueDate Must Must = BeginDate Where Sample OR ValueText is Null AND F_SYSTEM >1 and NHS in (1;2;3;4;5;6;7;8;9)
        #Assuming ValueDate must be >= BeginDate otherwise all items would fail
        print("Running rule SJF78...")
        self.df['SJF78'] = True
        tempDF = self.df.copy() 
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(pd.to_datetime(tempDF['IRI_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['IRI_VALUE_DATE'].isna()]
        self.df['SJF78'].iloc[tempDF.index.tolist()] = False 

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] > 1]
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[(pd.to_datetime(tempDF['IRI_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['IRI_VALUE_DATE'].isna()]
        self.df['SJF78'].iloc[tempDF.index.tolist()] = False 

    def sjf79(self):
        #RUTTING
        #ValueDate Must Must = BeginDate Where Sample OR ValueText is Null AND F_SYSTEM >1 and NHS in (1;2;3;4;5;6;7;8;9)
        #Assuming ValueDate must be >= BeginDate otherwise all items would fail
        print("Running rule SJF79...")
        self.df['SJF79'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(pd.to_datetime(tempDF['RUTTING_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['RUTTING_VALUE_DATE'].isna()]
        self.df['SJF79'].iloc[tempDF.index.tolist()] = False  

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] > 1]
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[(pd.to_datetime(tempDF['RUTTING_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['RUTTING_VALUE_DATE'].isna()]
        self.df['SJF79'].iloc[tempDF.index.tolist()] = False  



    def sjf80(self):
        #FAULTING
        #ValueDate Must Must = BeginDate Where Sample OR ValueText is Null AND F_SYSTEM >1 and NHS in (1;2;3;4;5;6;7;8;9)
        #Assuming ValueDate must be >= BeginDate otherwise all items would fail
        print("Running rule SJF80...")
        self.df['SJF80'] = True
        tempDF = self.df.copy() 
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(pd.to_datetime(tempDF['FAULTING_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['FAULTING_VALUE_DATE'].isna()]
        self.df['SJF80'].iloc[tempDF.index.tolist()] = False  

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] > 1 ]
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[(pd.to_datetime(tempDF['FAULTING_VALUE_DATE']) < tempDF['BEGIN_DATE']) | tempDF['FAULTING_VALUE_DATE'].isna()]
        self.df['SJF80'].iloc[tempDF.index.tolist()] = False  


    def sjf81(self):
        #CRACKING_PERCENT
        #ValueDate Must Must >= (BeginDate – 1) Where Sample OR ValueText is Null AND F_SYSTEM >1 and NHS in (1;2;3;4;5;6;7;8;9)
        print("Running rule SJF81...")
        self.df['SJF81'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        tempDF = tempDF[(pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE']) < (tempDF['BEGIN_DATE'] - pd.DateOffset(years=1))) | tempDF['CRACKING_PERCENT_VALUE_DATE'].isna()]
        self.df['SJF81'].iloc[tempDF.index.tolist()] = False

        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] > 1]
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[(pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE']) < (tempDF['BEGIN_DATE'] - pd.DateOffset(years=1))) | tempDF['CRACKING_PERCENT_VALUE_DATE'].isna()]
        self.df['SJF81'].iloc[tempDF.index.tolist()] = False

    def sjf82a(self):
        #Cracking Percent
        #ValueDate Must = BeginDate  Where ValueText is Null AND F_SYSTEM =1
        print("Running rule SJF82a...")
        self.df['SJF82a'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE']).dt.year != tempDF['BEGIN_DATE'].dt.year]
        self.df['SJF82a'].iloc[tempDF.index.tolist()] = False


    def sjf82b(self):
        #Faulting
        #ValueDate Must = BeginDate  Where ValueText is Null AND F_SYSTEM =1
        print("Running rule SJF82b...")
        self.df['SJF82b'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['FAULTING_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[pd.to_datetime(tempDF['FAULTING_VALUE_DATE']).dt.year != tempDF['BEGIN_DATE'].dt.year]
        self.df['SJF82b'].iloc[tempDF.index.tolist()] = False

    def sjf82c(self):
        #IRI
        #ValueDate Must = BeginDate  Where ValueText is Null AND F_SYSTEM =1
        print("Running rule SJF82c...")
        self.df['SJF82c'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['IRI_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[pd.to_datetime(tempDF['IRI_VALUE_DATE']).dt.year != tempDF['BEGIN_DATE'].dt.year]
        self.df['SJF82c'].iloc[tempDF.index.tolist()] = False

    def sjf82d(self):
        #Rutting
        #ValueDate Must = BeginDate  Where ValueText is Null AND F_SYSTEM =1
        print("Running rule SJF82d...")
        self.df['SJF82d'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING_VALUE_TEXT'].isna()]
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[pd.to_datetime(tempDF['RUTTING_VALUE_DATE']).dt.year != tempDF['BEGIN_DATE'].dt.year]
        self.df['SJF82d'].iloc[tempDF.index.tolist()] = False

    def sjf83a(self):
        #CRACKING_PERCENT
        #ValueText Must Be In (A;B;C;D;E) Where ValueDate <> BeginDate and F_SYSTEM = 1 OR if ValueDate < BeginDate -1 on NHS
        #Rule not created as we don't have the ValueText information
        print("Running rule SJF83a...")
        self.df['SJF83a'] = True
    
    def sjf83b(self):
        #Faulting
        #ValueText Must Be In (A;B;C;D;E) Where ValueDate <> BeginDate and F_SYSTEM = 1 OR if ValueDate < BeginDate -1 on NHS
        #Rule not created as we don't have the ValueText information
        print("Running rule SJF83b...")
        self.df['SJF83b'] = True

    def sjf83c(self):
        #IRI
        #ValueText Must Be In (A;B;C;D;E) Where ValueDate <> BeginDate and F_SYSTEM = 1 OR if ValueDate < BeginDate -1 on NHS
        #Rule not created as we don't have the ValueText information
        print("Running rule SJF83c...")
        self.df['SJF83c'] = True

    def sjf83d(self):
        #Rutting
        #ValueText Must Be In (A;B;C;D;E) Where ValueDate <> BeginDate and F_SYSTEM = 1 OR if ValueDate < BeginDate -1 on NHS
        #Rule not created as we don't have the ValueText information
        print("Running rule SJF83d...")
        self.df['SJF83d'] = True

    def sjf84(self):
        #PSR ValueDate Must Must >= BeginDate – 1 Where Sample OR F_SYSTEM >1 and NHS in (1;2;3;4;5;6;7;8;9)
        #Rule not implemented as we don't report PSR
        print("Running rule SJF84...")
        self.df['SJF84'] = True


    def sjf85(self):
        #ValueDate Must = BeginDate Where PSR ValueText is "A" AND F_SYSTEM =1
        #Rule not implemented as we don't report PSR
        print("Running rule SJF85...")
        self.df['SJF85'] = True

    def sjf86(self):
        #Where F_SYSTEM =1; and IRI is Null; PSR ValueNumeric Must be >0 and PSR ValueText must = A
        #Rule not implemented as we don't report PSR
        print("Running rule SJF86...")
        self.df['SJF86'] = True

    def sjf87(self):
        #RUTTING < 1
        print("Running rule SJF87...")
        self.df['SJF87'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['RUTTING'] >= 1]
        self.df['SJF87'].iloc[tempDF.index.tolist()] = False

    def sjf88(self):
        #FAULTING<=1
        print("Running rule SJF88...")
        self.df['SJF88'] = True
        tempDF = self.df.copy() 
        tempDF = tempDF[tempDF['FAULTING'] > 1]
        self.df['SJF88'].iloc[tempDF.index.tolist()] = False 


    def sjf89(self):
        #Where SURFACE_TYPE is in (2;6;7;8) CRACKING_PERCENT <= X based on LANE_WIDTH. Where X = Max % AC Cracking on AC Cracking Validation table.
        print("Running rule SJF89...")
        self.df['SJF89'] = True
        
        #Lane Width : Max % Cracking
        """
        Lane width values 6.0, 7.0, 17.0, and 18.0 were missing from AC validation table, however lane_widths of 6,7,17 and 18 are valid.
        Made max_cracking_percent 100% for 6,7 and 40% for 17,18. Needs to be updated if we find out correct values for these
        """
        acValidation = {
            6.0 : 100.00, #NOT ON AC VALIDATION TABLE
            7.0 : 100.00, #NOT ON AC VALIDATION TABLE
            8.0 : 81.30,
            9.0 : 72.20,
            10.0 : 65.00,
            11.0 : 59.10,
            12.0 : 54.20,
            13.0 : 50.00,
            14.0 : 46.40,
            15.0 : 43.30,
            16.0 : 40.60,
            17.0 : 40.00, #NOT ON AC VALIDATION TABLE
            18.0 : 40.00  #NOT ON AC VALIDATION TABLE
        }


        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].notna()]
        tempDF['MAX_CRACKING'] = tempDF['LANE_WIDTH'].copy()
        tempDF['MAX_CRACKING'].replace(acValidation, inplace=True)
        tempDF = tempDF[tempDF['CRACKING_PERCENT'] > tempDF['MAX_CRACKING']]
        self.df['SJF89'].iloc[tempDF.index.tolist()] = False 

    def sjf90(self):
        #Where SURFACE_TYPE is in (3;4;5;9;10) CRACKING_PERCENT < .75
        print("Running rule SJF90...")
        self.df['SJF90'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,5,9,10])]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'] >= 75]
        self.df['SJF90'].iloc[tempDF.index.tolist()] = False

    def sjf91(self):
        #ValueDate <= BeginDate
        print("Running rule SJF91...")
        self.df['SJF91'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['YEAR_LAST_CONSTRUCTION_VALUE_DATE'] > tempDF['BEGIN_DATE'].dt.year]
        self.df['SJF91'].iloc[tempDF.index.tolist()] = False

    def sjf92(self):
        #THICKNESS_RIGID must be Null WHERE SURFACE_TYPE in (2;6) 
        print("Running rule SJF92...")
        self.df['SJF92'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6])]
        tempDF = tempDF[tempDF['THICKNESS_RIGID'].notna()]
        self.df['SJF92'].iloc[tempDF.index.tolist()] = False


    def sjf93(self):
        #THICKNESS_FLEXIBLE must be Null WHERE SURFACE_TYPE in (3;4;5;9;10)
        print("Running rule SJF93...")
        self.df['SJF93'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,5,9,10])]
        tempDF = tempDF[tempDF['THICKNESS_FLEXIBLE'].notna()]
        self.df['SJF93'].iloc[tempDF.index.tolist()] = False

    def sjf94(self):
        #THICKNESS_FLEXIBLE MUST NOT be Null WHERE SURFACE_TYPE IN (7;8)
        print("Running rule SJF94...")
        self.df['SJF94'] = True
        tempDF = self.df.copy() 
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([7,8])]
        tempDF = tempDF[tempDF['THICKNESS_FLEXIBLE'].isna()]
        self.df['SJF94'].iloc[tempDF.index.tolist()] = False

    def sjf95(self):
        #THICKNESS_RIGID MUST NOT be Null WHERE SURFACE_TYPE IN (7;8)
        print("Running rule SJF95...")
        self.df['SJF95'] = True
        tempDF = self.df.copy() 
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([7,8])]
        tempDF = tempDF[tempDF['THICKNESS_RIGID'].isna()]
        self.df['SJF95'].iloc[tempDF.index.tolist()] = False

    def sjf96(self):
        #DIR_THROUGH_LANES ValueNumeric Must be < OR = ValueNumeric for THROUGH_LANES
        print("Running rule SJF96...")
        self.df['SJF96'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'] > tempDF['THROUGH_LANES']]
        self.df['SJF96'].iloc[tempDF.index.tolist()] = False


    def sjf97(self):
        #IF TRAVEL_TIME_CODE is reported; it must cover NHS
        #We don't report TRAVEL_TIME
        print("Running rule SJF97...")
        self.df['SJF97'] = True

    def sjf98(self):
        #MAINTENANCE_OPERATIONS ValueNumeric <> OWNERSHIP ValueNumeric
        #Treating this rule as "MAINTENANCE_OPERATIONS must equal OWNERSHIP"
        print("Running rule SJF98...")
        self.df['SJF98'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['MAINTENANCE_OPERATIONS'].notna() | tempDF['OWNERSHIP'].notna()]
        tempDF = tempDF[tempDF['MAINTENANCE_OPERATIONS'] != tempDF['OWNERSHIP']] 
        self.df['SJF98'].iloc[tempDF.index.tolist()] = False

    def sjf99(self):
        #Sample crosses TOPS.  
        # The extent of a given Sample Panel Section extends beyond the extent of the associated TOPS section.  
        # Samples should match the length of TOPS sections or be shorter; but can not be longer.
        print("Running rule SJF99...")
        self.df['SJF99'] = True
   

    def sjf100(self):
        #Samples may only exist where Facility_Type IN 1;2 and (F_System = 1-5 or F_System = 6 and Urban Code <99999)
        print("Running rule SJF100...")
        self.df['SJF100'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[~tempDF['FACILITY_TYPE'].isin([1,2]) | ~tempDF['F_SYSTEM'].isin(range(1,6))]
        tempDF = tempDF[~tempDF['FACILITY_TYPE'].isin([1,2]) | (tempDF['F_SYSTEM'] != 6) | (tempDF['URBAN_CODE'].astype(float) >= 99999)]
        tempDF = tempDF[tempDF['HPMS_SAMPLE_NO'].notna()]
        self.df['SJF100'].iloc[tempDF.index.tolist()] = False 
    
    def run(self):
        self.sjf01()
        self.sjf02()
        self.sjf03()
        self.sjf04()
        self.sjf05()
        self.sjf06()
        self.sjf07()
        self.sjf08()
        self.sjf09()
        self.sjf10()
        self.sjf11()
        self.sjf12()
        self.sjf13()
        self.sjf14()
        self.sjf15()
        self.sjf16()
        self.sjf17()
        self.sjf18()
        self.sjf19()
        self.sjf20()
        self.sjf21()
        self.sjf22()
        self.sjf23()
        self.sjf24()
        self.sjf25()
        self.sjf26()
        self.sjf27()
        self.sjf28()
        self.sjf29()
        self.sjf30()
        self.sjf31()
        self.sjf32()
        self.sjf33()
        self.sjf34()
        self.sjf35()
        self.sjf36()
        self.sjf37()
        self.sjf38()
        self.sjf39()
        self.sjf40()
        self.sjf41()
        self.sjf42()
        self.sjf43()
        self.sjf44()
        self.sjf45()
        self.sjf46()
        self.sjf47()
        self.sjf48()
        self.sjf49()
        self.sjf50()
        self.sjf51()
        self.sjf52()
        self.sjf53()
        self.sjf54()
        self.sjf55()
        self.sjf56()
        self.sjf57()
        self.sjf58()
        self.sjf59()
        self.sjf60()
        self.sjf61()
        self.sjf62()
        self.sjf63()
        self.sjf64()
        self.sjf65()
        self.sjf66()
        self.sjf67()
        self.sjf68()
        self.sjf69()
        self.sjf70()
        self.sjf71()
        self.sjf72()
        self.sjf73()
        self.sjf74()
        self.sjf75()
        self.sjf76()
        self.sjf77()
        self.sjf78()
        self.sjf79()
        self.sjf80()
        self.sjf81()
        self.sjf82a()
        self.sjf82b()
        self.sjf82c()
        self.sjf82d()
        self.sjf83a()
        self.sjf83b()
        self.sjf83c()
        self.sjf83d()
        self.sjf84()
        self.sjf85()
        self.sjf86()
        self.sjf87()
        self.sjf88()
        self.sjf89()
        self.sjf90()
        self.sjf91()
        self.sjf92()
        self.sjf93()
        self.sjf94()
        self.sjf95()
        self.sjf96()
        self.sjf96()
        self.sjf97()
        self.sjf98()
        self.sjf99()
        self.sjf100()
    
    def create_output(self, template='fullSpatialErrors_template.xlsx', outfilename='rules_summary.xlsx'):
        #Reads sheet on template that list all data items associated with each rule and converts to dictionary
        dataItemsDF = pd.read_excel(template, sheet_name="ruleDataItems", usecols='A,B', nrows=106)
        dataItemsDF['Rule'] = dataItemsDF['Rule'].str.replace("-", "")
        dataItemsDF['Data_Items'] = dataItemsDF['Data_Items'].str.split(",")
        ruleDict = dict(zip(dataItemsDF['Rule'], dataItemsDF['Data_Items']))
        fwhaRuleDict = ruleDict.copy()
        for rule in ruleDict.keys():
            fwhaRuleDict[rule+"_FHWA"] = fwhaRuleDict.pop(rule)


        #Reads the rule descripts off of summary sheet and converts to dictionary
        ruleDescDF = pd.read_excel(template, sheet_name="Summary", usecols="A,D")
        ruleDesc = dict(zip(ruleDescDF['Rule'], ruleDescDF['Description']))

        #Create copy of template to write to
        shutil.copy(template, outfilename)

        with pd.ExcelWriter(outfilename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            tempDF = self.df.copy()
            numFailed = []
            numPassed = []
            lenFailed = []
            fwhaFailed = []

            #Get counts for failed/passed/length of failed
            print("Updating summary sheet on",outfilename,"...")
            for rule in ruleDict.keys():
                #Assumes all passes for rules not ran
                try:
                    numFailed.append(tempDF[tempDF[rule]==False].shape[0])
                    numPassed.append(tempDF[tempDF[rule]==True].shape[0])
                    lenFailed.append((tempDF[tempDF[rule]==False]['EMP'] - tempDF[tempDF[rule]==False]['BMP']).sum())
                except KeyError:
                    numFailed.append(0)
                    numPassed.append(self.df.shape[0])
                    lenFailed.append(0)

            for rule in fwhaRuleDict.keys():
                try:
                    fwhaFailed.append(tempDF[tempDF[rule]==False].shape[0])
                except KeyError:
                    fwhaFailed.append(0)

            failedDF = pd.DataFrame(numFailed)
            passedDF = pd.DataFrame(numPassed)
            lengthDF = pd.DataFrame(lenFailed)
            fwhaFailedDF = pd.DataFrame(fwhaFailed)

            #Write counts to Summary sheet
            failedDF.to_excel(writer, sheet_name='Summary', startcol=4, startrow=1, header=False, index=False)
            passedDF.to_excel(writer, sheet_name='Summary', startcol=5, startrow=1, header=False, index=False)
            lengthDF.to_excel(writer, sheet_name='Summary', startcol=6, startrow=1, header=False, index=False)
            fwhaFailedDF.to_excel(writer, sheet_name='Summary', startcol=7, startrow=1, header=False, index=False)

            #Create sheets for each rule containing all failed rows (using only columns that the specific rule references)
            for rule in ruleDict.keys():
                tempDF = self.df.copy()
                #Checks to make sure rule has data items associated with it (will be a list if dataItems exists, otherwise will be float (np.nan))
                if type(ruleDict[rule])==list:
                    #Tries using RULENAME (i.e. SJF01) in dataset which is added if the rule is ran
                    #If rule is not run, no column will exist with the rulename, catches KeyError and prints message.
                    try:
                        if tempDF[tempDF[rule]==False].shape[0] > 0:
                            print("Creating error sheet for rule:",rule)
                            dataItems = ['RouteID', 'BMP', 'EMP']
                            [dataItems.append(x) for x in ruleDict[rule] if x not in dataItems]
                            tempDF = tempDF[tempDF[rule]==False]
                            tempDF = tempDF[dataItems]
                            tempDF.to_excel(writer, sheet_name=rule, startrow=1, index=False)
                            worksheet = writer.sheets[rule]
                            worksheet['A1'] = f'=HYPERLINK("#Summary!A1", "Summary Worksheet")'
                            worksheet['A1'].font = Font(underline='single', color='0000EE')
                            #Autofit columns
                            for i in range(1, worksheet.max_column+1):
                                worksheet.column_dimensions[get_column_letter(i)].width = 20
                            #Add rule description to sheet
                            worksheet['B1'] = ruleDesc[rule]

                        else:
                            print("No failed rows for rule:",rule)

                    except KeyError:
                        print(rule,"not found in DF. Sheet was not created in rules_summary.xlsx")
                else:
                    print("No data items for rule",rule,", Sheet not created.")

    


df = pd.read_csv('all_submission_data.csv',dtype={'URBAN_CODE':str, 'SampleId':str})

c = full_spatial_functions(df)  
# c.run()
c.run()
c.create_output()
print(c.df)
# c.df.to_csv('test_functions_matt_sucks.csv', index=False)