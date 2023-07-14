import pandas as pd
import shutil
import datetime
import warnings
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

class pm2_validations():

    def __init__(self, df):
        self.df = df
        #Setting Pav_Rep_Method as 2 since that's the value for all entries for the forseeable future. This will need to be updated if it changes
        try:
            self.df['Pav_Rep_Method']
        except KeyError:
            self.df['Pav_Rep_Method'] = 2

        try:
            self.df['BEGIN_DATE']
        except KeyError:
            self.df['BEGIN_DATE'] = datetime.datetime(2022,1,1)

        self.df.rename(columns = {
            "FsystemVn":"F_SYSTEM",
            "NhsVn":"NHS",
            "NnVn":"NN",
            "UrbanIdVn":"URBAN_CODE",
            "FacilityTypeVn":"FACILITY_TYPE",
            "StructureTypeVn":"STRUCTURE_TYPE",
            "OwnershipVn":"OWNERSHIP",
            "CountyIdVn":"COUNTY_ID",
            "MaintenanceOperationsVn":"MAINTENANCE_OPERATIONS",
            "ThroughLanesVn":"THROUGH_LANES",
            "PeakLanesVn":"PEAK_LANES",
            "CounterPeakLanesVn":"COUNTER_PEAK_LANES",
            "LaneWidthVn":"LANE_WIDTH",
            "MedianTypeVn":"MEDIAN_TYPE",
            "MedianWidthVn":"MEDIAN_WIDTH",
            "ShoulderTypeVn":"SHOULDER_TYPE",
            "ShoulderWidthRVn":"SHOULDER_WIDTH_R",
            "ShoulderWidthLVn":"SHOULDER_WIDTH_L",
            "PeakParkingVn":"PEAK_PARKING",
            "DirThroughLanesVn":"DIR_THROUGH_LANES",
            "TurnLanesRVn":"TURN_LANES_R",
            "TurnLanesLVn":"TURN_LANES_L",
            "SignalTypeVn":"SIGNAL_TYPE",
            "PctGreenTimeVn":"PCT_GREEN_TIME",
            "NumberSignalsVn":"NUMBER_SIGNALS",
            "StopSignsVn":"STOP_SIGNS",
            "AtGradeOtherVn":"AT_GRADE_OTHER",
            "AadtVn":"AADT",
            "AadtVt":"AADT_VALUE_TEXT",
            "AadtVd":"AADT_VALUE_DATE",
            "AadtsingleUnitVn":"AADT_SINGLE_UNIT",
            "AadtsingleUnitVt":"AADT_SINGLE_UNIT_VALUE_TEXT",
            "AadtsingleUnitVd":"AADT_SINGLE_UNIT_VALUE_DATE",
            "AadtcombinationVn":"AADT_COMBINATION",
            "PctdhsingleVn":"PCT_DH_SINGLE_UNIT",
            "PctdhcombinationVn":"PCT_DH_COMBINATION",
            "KfactorVn":"K_FACTOR",
            "DirFactorVn":"DIR_FACTOR",
            "FutureAadtVn":"FUTURE_AADT",
            "AccessControlVn":"ACCESS_CONTROL",
            "SpeedLimitVn":"SPEED_LIMIT",
            "IriVn":"IRI",
            "IriVt":"IRI_VALUE_TEXT",
            "IriVd":"IRI_VALUE_DATE",
            "SurfaceTypeVn":"SURFACE_TYPE",
            "RuttingVn":"RUTTING",
            "RuttingVt":"RUTTING_VALUE_TEXT",
            "RuttingVd":"RUTTING_VALUE_DATE",
            "FaultingVn":"FAULTING",
            "FaultingVt":"FAULTING_VALUE_TEXT",
            "FaultingVd":"FAULTING_VALUE_DATE",
            "CrackingPercentVn":"CRACKING_PERCENT",
            "CrackingPercentVt":"CRACKING_PERCENT_VALUE_TEXT",
            "CrackingPercentVd":"CRACKING_PERCENT_VALUE_DATE",
            "YearLastImprovementVd":"YEAR_LAST_IMPROVEMENT_VALUE_DATE",
            "YearLastConstructionVd":"YEAR_LAST_CONSTRUCTION_VALUE_DATE",
            "LastOverlayThicknessVn":"LAST_OVERLAY_THICKNESS",
            "ThicknessRigidVn":"THICKNESS_RIGID",
            "ThicknessFlexibleVn":"THICKNESS_FLEXIBLE",
            "BaseTypeVn":"BASE_TYPE",
            "BaseThicknessVn":"BASE_THICKNESS",
            "SoilTypeVn":"SOIL_TYPE",
            "WideningPotentialVn":"WIDENING_POTENTIAL",
            "CurvesAVn":"CURVES_A",
            "CurvesBVn":"CURVES_B",
            "CurvesCVn":"CURVES_C",
            "CurvesDVn":"CURVES_D",
            "CurvesEVn":"CURVES_E",
            "CurvesFVn":"CURVES_F",
            "TerrarinTypeVn":"TERRAIN_TYPE",
            "GradesAVn":"GRADES_A",
            "GradesBVn":"GRADES_B",
            "GradesCVn":"GRADES_C",
            "GradesDVn":"GRADES_D",
            "GradesEVn":"GRADES_E",
            "GradesFVn":"GRADES_F",
            "PctPassSightVn":"PCT_PASS_SIGHT",
            "RouteQualifier":"ROUTE_QUALIFIER",
            "RouteSigning":"ROUTE_SIGNING",
            "RouteName":"ROUTE_NAME_VALUE_TEXT",
            "RouteNumber":"ROUTE_NUMBER",
            "SampleId":"HPMS_SAMPLE_NO"
        }, inplace=True)
        



    def SJPM201(self):
        #Dir_Through_Lanes must exist where F_System = 1 and Facility_Type in (1;2;6) 
        # and RoadEventCollectionMethods Pav_Rep_Method = 2

        print("Running rule SJPM201")
        self.df['SJPM201'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['DIR_THROUGH_LANES'].isna()]
        self.df['SJPM201'].iloc[tempDF.index.tolist()] = False

    def SJPM202(self):
        #Through_Lanes must exist where F_System = 1 and Facility_Type in (1;2)

        print("Running rule SJPM202")
        self.df['SJPM202'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['THROUGH_LANES'].isna()]
        self.df['SJPM202'].iloc[tempDF.index.tolist()] = False

    def SJPM203(self):
        #Through_Lanes must exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2)

        print("Running rule SJPM203")
        self.df['SJPM203'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['THROUGH_LANES'].isna()]
        self.df['SJPM203'].iloc[tempDF.index.tolist()] = False

    def SJPM204(self):
        #Surface Type should exist where F_System = 1 and Facility_Type in (1;2) 
        # and RoadEventCollectionMethod Pav_Rep_Method =1

        print("Running rule SJPM204")
        self.df['SJPM204'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 1]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isna()]
        self.df['SJPM204'].iloc[tempDF.index.tolist()] = False

    def SJPM205(self):
        #Surface Type should exist where F_System = 1 and Facility_Type in (1;2;6) 
        # and RoadEventCollectionMethod.PavRepMethod =2

        print("Running rule SJPM205")
        self.df['SJPM205'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isna()]
        self.df['SJPM205'].iloc[tempDF.index.tolist()] = False

    def SJPM206(self):
        #Surface Type should exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2)

        print("Running rule SJPM206")
        self.df['SJPM206'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isna()]
        self.df['SJPM206'].iloc[tempDF.index.tolist()] = False       

    def SJPM207(self):
        #IRI should exist where F_SYSTEM = 1 and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and IRI_Date is < RecordDateYear-1

        print("Running rule SJPM207")
        self.df['SJPM207'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE']!=1]
        tempDF['IRI_VALUE_DATE'] = pd.to_datetime(tempDF['IRI_VALUE_DATE'])
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['SJPM207'].iloc[tempDF.index.tolist()] = False

    def SJPM208(self):
        #IRI should exist where F_SYSTEM = 1 and Facility_Type in (1;2;6) 
        # and Structure_Type is not 1 and PSR is Null and IRI_Date is < RecordDateYear-1 and RoadEventCollectionMethod.PavRepMethod =2

        print("Running rule SJPM208")
        self.df['SJPM208'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE']!=1]
        tempDF['IRI_VALUE_DATE'] = pd.to_datetime(tempDF['IRI_VALUE_DATE'])
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['SJPM208'].iloc[tempDF.index.tolist()] = False

    def SJPM209(self):
        #IRI should exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and IRI_Date is < RecordDateYear-2

        print("Running rule SJPM209")
        self.df['SJPM209'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE']!=1]
        tempDF['IRI_VALUE_DATE'] = pd.to_datetime(tempDF['IRI_VALUE_DATE'])
        tempDF = tempDF[tempDF['IRI_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year-2)]
        tempDF = tempDF[tempDF['IRI'].isna()]
        self.df['SJPM209'].iloc[tempDF.index.tolist()] = False



    def SJPM210(self):
        #Cracking Percent should exist where F_SYSTEM = 1 and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and Cracking Percent_Date is < RecordDateYear-1

        print("Running rule SJPM210")
        self.df['SJPM210'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM']==1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] !=1]
        tempDF['CRACKING_PERCENT_VALUE_DATE'] = pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE'])
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['SJPM210'].iloc[tempDF.index.tolist()] = False

    def SJPM211(self):
        #Cracking Percent should exist where F_SYSTEM = 1 and Facility_Type in (1;2;6) and Structure_Type is not 1 and PSR is Null 
        # and Cracking Percent_Date is < RecordDateYear-1 and RoadEventCollectionMethod.PavRepMethod =2

        print("Running rule SJPM211")
        self.df['SJPM211'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['CRACKING_PERCENT_VALUE_DATE'] = pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE'])
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['SJPM211'].iloc[tempDF.index.tolist()] = False



    def SJPM212(self):
        #Cracking Percent should exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and Cracking Percent_Date is < RecordDateYear-2

        print("Running rule SJPM212")
        self.df['SJPM212'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['CRACKING_PERCENT_VALUE_DATE'] = pd.to_datetime(tempDF['CRACKING_PERCENT_VALUE_DATE'])
        tempDF = tempDF[tempDF['CRACKING_PERCENT_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 2)]
        tempDF = tempDF[tempDF['CRACKING_PERCENT'].isna()]
        self.df['SJPM212'].iloc[tempDF.index.tolist()] = False

    def SJPM213(self):
        #Faulting should exist where F_SYSTEM = 1 and Facility_Type in (1;2) and Structure_Type is not 1 and PSR is Null 
        # and Faulting_Date is < RecordDateYear-1 and Surface_Type in (3;4;9;10)

        print("Running rule SJPM213")
        self.df['SJPM213'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['FAULTING_VALUE_DATE'] = pd.to_datetime(tempDF['FAULTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,9,10])]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['SJPM213'].iloc[tempDF.index.tolist()] = False

    def SJPM214(self):
        #Faulting should exist where F_SYSTEM = 1 and Facility_Type in (1;2;6) and Structure_Type is not 1 and PSR is Null 
        # and Faulting_Date is < RecordDateYear-1 and RoadEventCollectionMethod.PavRepMethod =2 and Surface_Type in (3;4;9;10)

        print("Running rule SJPM214")
        self.df['SJPM214'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['FAULTING_VALUE_DATE'] = pd.to_datetime(tempDF['FAULTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,9,10])]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['SJPM214'].iloc[tempDF.index.tolist()] = False

    def SJPM215(self):
        #Faulting should exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and Faulting_Date is < RecordDateYear-2 and Surface_Type in (3;4;9;10)

        print("Running rule SJPM215")
        self.df['SJPM215'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['FAULTING_VALUE_DATE'] = pd.to_datetime(tempDF['FAULTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['FAULTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year -2 )]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([3,4,9,10])]
        tempDF = tempDF[tempDF['FAULTING'].isna()]
        self.df['SJPM215'].iloc[tempDF.index.tolist()] = False


    def SJPM216(self):
        #Rutting should exist where F_SYSTEM = 1 and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and Rutting_Date is < RecordDateYear-1 and Surface_Type in (2;6;7;8)

        print("Running rule SJPM216")
        self.df['SJPM216'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['RUTTING_VALUE_DATE'] = pd.to_datetime(tempDF['RUTTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['SJPM216'].iloc[tempDF.index.tolist()] = False

    def SJPM217(self):
        #Rutting should exist where F_SYSTEM = 1 and Facility_Type in (1;2;6) and Structure_Type is not 1 and PSR is Null 
        # and Rutting_Date is < RecordDateYear-1 and RoadEventCollectionMethod.PavRepMethod =2 and Surface_Type in (2;6;7;8)

        print("Running rule SJPM217")
        self.df['SJPM217'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['F_SYSTEM'] == 1]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2,6])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['RUTTING_VALUE_DATE'] = pd.to_datetime(tempDF['RUTTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 1)]
        tempDF = tempDF[tempDF['Pav_Rep_Method'] == 2]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['SJPM217'].iloc[tempDF.index.tolist()] = False

    def SJPM218(self):
        #Rutting should exist where NHS in (1;2;3;4;5;6;7;8;9) and Facility_Type in (1;2) 
        # and Structure_Type is not 1 and PSR is Null and Rutting_Date is < RecordDateYear-2 and Surface_Type in (2;6;7;8)

        print("Running rule SJPM218")
        self.df['SJPM218'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[tempDF['NHS'].isin(range(1,10))]
        tempDF = tempDF[tempDF['FACILITY_TYPE'].isin([1,2])]
        tempDF = tempDF[tempDF['STRUCTURE_TYPE'] != 1]
        tempDF['RUTTING_VALUE_DATE'] = pd.to_datetime(tempDF['RUTTING_VALUE_DATE'])
        tempDF = tempDF[tempDF['RUTTING_VALUE_DATE'].dt.year < (tempDF['BEGIN_DATE'].dt.year - 2)]
        tempDF = tempDF[tempDF['SURFACE_TYPE'].isin([2,6,7,8])]
        tempDF = tempDF[tempDF['RUTTING'].isna()]
        self.df['SJPM218'].iloc[tempDF.index.tolist()] = False

    def SJPM219(self):
        #IRI values < 30 and > 400 should be reviewed and where valid; 
        # explanation provided in submission comments. 

        print("Running rule SJPM219")
        self.df['SJPM219'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['IRI'] < 30)| (tempDF['IRI'] > 400)]
        self.df['SJPM219'].iloc[tempDF.index.tolist()] = False

    def SJPM230(self):
        #Cracking_Percent event segments must align with with the Begin and End points for IRI segments.

        print("Running rule SJPM230")
        self.df['SJPM230'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['CRACKING_PERCENT'].notna() & tempDF['IRI'].isna())]
        self.df['SJPM230'].iloc[tempDF.index.tolist()] = False

    def SJPM231(self):
        #Rutting event segments must align with with the Begin and End points for IRI segments.

        print("Running rule SJPM231")
        self.df['SJPM231'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['RUTTING'].notna() & tempDF['IRI'].isna())]
        self.df['SJPM231'].iloc[tempDF.index.tolist()] = False

    def SJPM232(self):
        #Faulting event segments must align with with the Begin and End points for IRI segments.

        print("Running rule SJPM232")
        self.df['SJPM232'] = True
        tempDF = self.df.copy()
        tempDF = tempDF[(tempDF['FAULTING'].notna() & tempDF['IRI'].isna())]
        self.df['SJPM232'].iloc[tempDF.index.tolist()] = False



    def run(self):
        self.SJPM201()
        self.SJPM202()
        self.SJPM203()
        self.SJPM204()
        self.SJPM205()
        self.SJPM206()
        self.SJPM207()
        self.SJPM208()
        self.SJPM209()
        self.SJPM210()
        self.SJPM211()
        self.SJPM212()
        self.SJPM213()
        self.SJPM214()
        self.SJPM215()
        self.SJPM216()
        self.SJPM217()
        self.SJPM218()
        self.SJPM219()
        self.SJPM230()
        self.SJPM231()
        self.SJPM232()

    def create_output(self, template='pm2_rules_summary_template.xlsx', outfilename='pm2_rules_summary.xlsx'):
        #Reads sheet on template that list all data items associated with each rule and converts to dictionary
        dataItemsDF = pd.read_excel(template, sheet_name="ruleDataItems", usecols='A,B', nrows=23)
        dataItemsDF['Rule'] = dataItemsDF['Rule'].str.replace("-", "")
        dataItemsDF['Data_Items'] = dataItemsDF['Data_Items'].str.split(",")
        ruleDict = dict(zip(dataItemsDF['Rule'], dataItemsDF['Data_Items']))
        fwhaRuleDict = ruleDict.copy()
        for rule in ruleDict.keys():
            fwhaRuleDict[rule+"_FHWA"] = fwhaRuleDict.pop(rule)


        #Reads the rule descripts off of summary sheet and converts to dictionary
        ruleDescDF = pd.read_excel(template, sheet_name="Summary", usecols="A,D")
        ruleDesc = dict(zip(ruleDescDF['Rule'], ruleDescDF['Description']))

        #Create copy of template to write to
        shutil.copy(template, outfilename)

        with pd.ExcelWriter(outfilename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            tempDF = self.df.copy()
            numFailed = []
            numPassed = []
            lenFailed = []
            fwhaFailed = []

            #Get counts for failed/passed/length of failed
            print("Updating summary sheet on",outfilename,"...")
            for rule in ruleDict.keys():
                #Assumes all passes for rules not ran
                try:
                    numFailed.append(tempDF[tempDF[rule]==False].shape[0])
                    numPassed.append(tempDF[tempDF[rule]==True].shape[0])
                    lenFailed.append((tempDF[tempDF[rule]==False]['EMP'] - tempDF[tempDF[rule]==False]['BMP']).sum())
                except KeyError:
                    numFailed.append(0)
                    numPassed.append(self.df.shape[0])
                    lenFailed.append(0)

            for rule in fwhaRuleDict.keys():
                try:
                    fwhaFailed.append(tempDF[tempDF[rule]==False].shape[0])
                except KeyError:
                    fwhaFailed.append(0)

            failedDF = pd.DataFrame(numFailed)
            passedDF = pd.DataFrame(numPassed)
            lengthDF = pd.DataFrame(lenFailed)
            fwhaFailedDF = pd.DataFrame(fwhaFailed)

            #Write counts to Summary sheet
            failedDF.to_excel(writer, sheet_name='Summary', startcol=4, startrow=1, header=False, index=False)
            passedDF.to_excel(writer, sheet_name='Summary', startcol=5, startrow=1, header=False, index=False)
            lengthDF.to_excel(writer, sheet_name='Summary', startcol=6, startrow=1, header=False, index=False)
            fwhaFailedDF.to_excel(writer, sheet_name='Summary', startcol=7, startrow=1, header=False, index=False)

            #Create sheets for each rule containing all failed rows (using only columns that the specific rule references)
            for rule in ruleDict.keys():
                tempDF = self.df.copy()
                #Checks to make sure rule has data items associated with it (will be a list if dataItems exists, otherwise will be float (np.nan))
                if type(ruleDict[rule])==list:
                    #Tries using RULENAME (i.e. SJF01) in dataset which is added if the rule is ran
                    #If rule is not run, no column will exist with the rulename, catches KeyError and prints message.
                    try:
                        if tempDF[tempDF[rule]==False].shape[0] > 0:
                            print("Creating error sheet for rule:",rule)
                            dataItems = ['RouteID', 'BMP', 'EMP']
                            [dataItems.append(x) for x in ruleDict[rule] if x not in dataItems]
                            tempDF = tempDF[tempDF[rule]==False]
                            tempDF = tempDF[dataItems]
                            tempDF.to_excel(writer, sheet_name=rule, startrow=1, index=False)
                            worksheet = writer.sheets[rule]
                            worksheet['A1'] = f'=HYPERLINK("#Summary!A1", "Summary Worksheet")'
                            worksheet['A1'].font = Font(underline='single', color='0000EE')
                            #Autofit columns
                            for i in range(1, worksheet.max_column+1):
                                worksheet.column_dimensions[get_column_letter(i)].width = 20
                            #Add rule description to sheet
                            worksheet['B1'] = ruleDesc[rule]

                        else:
                            print("No failed rows for rule:",rule)

                    except KeyError:
                        print(rule,"not found in DF. Sheet was not created in rules_summary.xlsx")
                else:
                    print("No data items for rule",rule,", Sheet not created.")



df = pd.read_csv(r'full_spat_with_error_columns.csv')

c = pm2_validations(df)
c.run()
c.create_output()