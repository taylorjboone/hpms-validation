import pandas as pd
import os
import geopandas as gpd
import shutil
import warnings
import wvdot_utils
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

cols_dict_old = {
    'BMP':'BeginPoint',
    'EMP':'EndPoint',
    '12_COUNTY':'COUNTY_ID',
    '2_AADT':'AADT',
    '2_FUTURE_AADT':'FUTURE_AADT',
    '3_ACCESS_CONTROL':'ACCESS_CONTROL',
    '4_ALT_ROUTE_NAME':'ALT_ROUTE_NAME',
    '5_AT_GRADE_OTHER':'AT_GRADE_OTHER',
    '6_AVE_LANE_WIDTH_FT':'LANE_WIDTH',
    '14_CRACKING_PERCENT':'CRACKING_PERCENT',
    '16_CURVE_CLASS':'CURVE_CLASSIFICATION',
    '17_DES_TRUCK_ROUTE':'NN',
    '20_FACILITY':'FACILITY_TYPE',
    '21_FAULTING':'FAULTING',
    '25_GRADE_CLASS':'GRADE_CLASSIFICATION',
    '30_IRI_VALUE':'IRI',
    '34_MEDIAN_WIDTH_FT':'MEDIAN_WIDTH',
    '35_NAT_FUNCTIONAL_CLASS':'FUNCTIONAL_CLASS',
    '36_NHS':'NHS',
    '37_NUMBER_SIGNALS':'NUMBER_SIGNALS',
    '39_OWNERSHIP':'OWNERSHIP',
    '41_PCT_GREEN_TIME':'PCT_GREEN_TIME',
    '42_PCT_PASS_SIGHT':'PCT_PASS_SIGHT',
    '43_PEAK_LANES':'PEAK_LANES',
    '43_COUNTER_PEAK_LANES':'COUNTER_PEAK_LANES',
    '45_PSR':'PSR',
    '52_RUTTING':'RUTTING',
    '56_SHOULDER_TYPE_RT':'SHOULDER_TYPE',
    '57_SHOULDER_WIDTH_LFT_FT':'SHOULDER_WIDTH_L',
    '58_SHOULDER_WIDTH_RT_FT':"SHOULDER_WIDTH_R",
    '59_SIGNAL_TYPE':'SIGNAL_TYPE',
    '63_SPEED_LIMIT_MPH':'SPEED_LIMIT',
    '66_STOP_SIGNS':'STOP_SIGNS',
    '70_SURFACE_TYPE':'SURFACE_TYPE',
    '71_TERRAIN_TYPE':'TERRAIN_TYPE',
    '74_NUM_THROUGH_LANES':'DIR_THROUGH_LANES_MAYBE',
    '75_TOLL_CHARGED':'TOLL_ID',
    '76_TOLL_TYPE':'TOLL_TYPE',
    '77_AADT_SINGLE':'AADT_SINGLE_UNIT',
    '77_AADT_COMBINATION':'AADT_COMBINATION',
    '77_PCT_PEAK_SINGLE':'PCT_DH_SINGLE_UNIT',
    '77_PCT_PEAK_COMBINATION':'PCT_DH_COMBINATION',
    '77_K_FACTOR':"K_FACTOR",
    '77_DIR_FACTOR':'DIR_FACTOR',
    '80_TURN_LANES_LFT':'TURN_LANES_L',
    '81_TURN_LANES_R':'TURN_LANES_R',
    '83_URBAN_CODE':'URBAN_ID',
    '84_WIDENING_OBSTACLE':'WIDENING_OBSTACLE',
    '85_WIDENING_POTENTIAL':'WIDENING_POTENTIAL',
    '115_STRAHNET':'STRAHNET'
}

cols_dict = {
    'URBAN_CODE':'URBAN_ID',
    'STRAHNET_TYPE': 'STRAHNET',
    'TRUCK':'NN',
}

def get_f_system(value):
    if value in [1, 11]:
        return int(1)
    elif value in [4, 12]:
        return int(2)
    elif value in [2, 14]:
        return int(3)
    elif value in [6, 16]:
        return int(4)
    elif value in [7, 17]:
        return int(5)
    elif value in [8, 18]:
        return int(6)
    elif value in [9, 19]:
        return int(7)

arnold = gpd.read_file(r'lrs_data/Dominant_Routes_2022.gdb')
arnold_rids = arnold['ROUTE_ID'].unique().tolist()

def load_defaults(df):
        # Standardizes columns names 
        df.rename(columns=cols_dict, inplace=True)
        cols = df.columns.tolist()

        if not 'Supp_Code' in cols:
            df['Supp_Code'] = df['RouteID'].str[9:11]

        # If ROUTE_NUMBER column is missing, adds and populates ROUTE_NUMBER pulled from RouteID
        if not 'ROUTE_NUMBER' in cols:
            df['ROUTE_NUMBER'] = df['RouteID'].str[3:7]
            df['ROUTE_NUMBER'] = df['ROUTE_NUMBER'].map(lambda x: x.lstrip('0'))
            print('Added ROUTE_NUMBER column')

        # If Sign System column is missing, adds and populates sign system pulled from RouteID
        if not 'ROUTE_SIGNING' in cols:
            df['ROUTE_SIGNING'] = df['RouteID'].str[2]
            print('Added ROUTE_SIGNING column')

        # Converts 1-19 F System to FHWA 1-7 F System
        if not 'F_SYSTEM' in cols:
            df['F_SYSTEM'] = df['FUNCTIONAL_CLASS'].map(lambda x: get_f_system(x))
            print('Added State Functional Class column')

        # If ROUTE_QUALIFIER column is missing, adds and populates ROUTE_QUALIFIER pulled from RouteID
        qualifier_dict = {'00':1,'01':2, '02':1, '03':5, '04':1, '05':1, '06':1, '07':1, '08':3, '09':3, '10':3, '11':3, '12':3, '13':9, '14':4, '15':6, '16':10, '17':10, '18':10, '19':10, '20':1, '21':10, '22':10, '23':10, '24':7, '25':10, '26':10, '27':10, '28':10, '51':10, '99':10}
        if not 'ROUTE_QUALIFIER' in cols:
            df['ROUTE_QUALIFIER'] = df['RouteID'].str[9:11]
            df['ROUTE_QUALIFIER'] = df['ROUTE_QUALIFIER'].map(lambda x: qualifier_dict[x])

        # Creates Dir through lanes from existing events
        if not 'Dir_Through_Lanes' in cols:
            df['Dir_Through_Lanes'] = ''
            df['Dir_Through_Lanes'].loc[((df['97_ORIG_SURVEY_DIRECTION'].notna()) & (df['97_ORIG_SURVEY_DIRECTION'] == '0'))] = df['PEAK_LANES']
            df['Dir_Through_Lanes'].loc[((df['97_ORIG_SURVEY_DIRECTION'].notna()) & (df['97_ORIG_SURVEY_DIRECTION'].isin(['1','A'])))] = df['COUNTER_PEAK_LANES']

        return df

class Cross_Validation():
    def __init__(self, df):
        # self.df = load_defaults(df)
        self.df = df.rename(columns=cols_dict)

    def check_geom(row):
        value = wvdot_utils.validate_geom.check_seg_valid(row['RouteID'],row['BMP'],row['EMP'])
        return value[-1]


    def inventory_spatial_join(self):
        df = self.df
        # Creates a column for each rule, outputs a False for each row that doesn't pass the rule for a column
        spatial_join_checks = {
            'sji01': ((df['F_SYSTEM'].notna()) & (df['RouteID'].isin(arnold_rids))),
            'sji02': ((df['FACILITY_TYPE'].notna()) & (df['RouteID'].isin(arnold_rids))),
            'sji03': ((df['OWNERSHIP'].notna()) & (df['RouteID'].isin(arnold_rids))),
            'sji04': ((df['URBAN_ID'].notna()) & (df['RouteID'].isin(arnold_rids))),
            'sji05': (((df['FACILITY_TYPE'].notna()) & (df['F_SYSTEM'].notna())) | df['FACILITY_TYPE'].isna()),
            'sji06': (((df['F_SYSTEM'].notna()) & (df['FACILITY_TYPE'].notna())) | df['F_SYSTEM'].isna()),
            'sji07': (((df['F_SYSTEM'] == 1) & (df['FACILITY_TYPE'].isin([1,2])) & (df['ROUTE_NUMBER'].notna())) | (df['F_SYSTEM'] != 1)),
            'sji08': (((df['F_SYSTEM'] == 1) & (df['NHS'] == 1)) | (df['F_SYSTEM'] != 1)),
            'sji09': (((df['ROUTE_NUMBER'].notna()) & (df['ROUTE_SIGNING'].notna())) | (df['ROUTE_NUMBER'].isna())),
            'sji10': (((df['ROUTE_NUMBER'].notna()) & (df['ROUTE_QUALIFIER'].notna())) | (df['ROUTE_NUMBER'].isna())),
            'sji11': (((df['F_SYSTEM'] == 1) & (df['STRAHNET'] == 1)) | (df['F_SYSTEM'] != 1)),
            'sji12': (((df['STRAHNET'].isin([1,2])) & (df['NHS'] == 1)) | (~df['STRAHNET'].isin([1,2]))),
            'sji13': (((df['F_SYSTEM'] == 1) & (df['NN'] == 1)) | (df['F_SYSTEM'] != 1))
        }

        tmp = df.copy()
        for k,v in spatial_join_checks.items():
            tmp[k] = v
        tmp2 = pd.DataFrame()
        for rule in spatial_join_checks.keys():
            tmp2 = pd.concat([tmp2, tmp[tmp[rule] == False]])
        return tmp2.drop_duplicates()


    def traffic_spatial_join(self):
        df = self.df
        spatial_join_checks = {
            'sjt01': ((df['AADT_SINGLE_UNIT'].isna()) | (df['AADT_SINGLE_UNIT'] < (0.4 * df['AADT']))),
            'sjt02': ((df['AADT_SINGLE_UNIT'].isna()) | (((df['AADT'] > 500) & (df['AADT_SINGLE_UNIT'] > 0)) | (df['AADT'] <= 500))),
            'sjt03': ((df['AADT_SINGLE_UNIT'].isna()) | ((df['AADT_SINGLE_UNIT'] + df['AADT_COMBINATION']) < (0.8 * df['AADT']))),
            'sjt04': ((df['AADT_SINGLE_UNIT'].isna()) | (((df['AADT_SINGLE_UNIT'] * 0.01) < (df['AADT'] * (df['PCT_DH_SINGLE_UNIT'] * .01))) & ((df['AADT'] * (df['PCT_DH_SINGLE_UNIT'] * .01)) < (df['AADT_SINGLE_UNIT'] * 0.5)))),
            'sjt05': ((df['PCT_DH_SINGLE_UNIT'].isna()) | ((df['PCT_DH_SINGLE_UNIT'] > 0) & (df['PCT_DH_SINGLE_UNIT'] < 25))),
            'sjt06': ((df['PCT_DH_SINGLE_UNIT'].isna()) | (((df['AADT_SINGLE_UNIT'] < 50) & (df['PCT_DH_SINGLE_UNIT'] == 0)) | (df['PCT_DH_SINGLE_UNIT'] != 0))),
            'sjt07': ((df['AADT_COMBINATION'].isna()) | (df['AADT_COMBINATION'] < (0.4 * df['AADT']))),
            'sjt08': ((df['AADT_COMBINATION'].isna()) | ((df['AADT'] <= 500) | ((df['AADT'] > 500) & (df['AADT_COMBINATION'] > 0)))),
            'sjt09': ((df['PCT_DH_COMBINATION'].isna()) | (((df['AADT_COMBINATION'] * .01) < (df['AADT'] * (df['PCT_DH_COMBINATION'] / 100))) & ((df['AADT'] * (df['PCT_DH_COMBINATION'] / 100)) < (df['AADT_COMBINATION'] * 0.5)))),
            'sjt10': ((df['PCT_DH_COMBINATION'].isna()) | ((df['PCT_DH_COMBINATION'] > 0) & (df['PCT_DH_COMBINATION'] < 25))),
            'sjt11': ((df['AADT_COMBINATION'].isna()) | ((df['PCT_DH_COMBINATION'] != 0) | ((df['PCT_DH_COMBINATION'] == 0) & (df['AADT_COMBINATION'] < 50)))),
            'sjt12': ((df['K_FACTOR'].isna()) | ((df['K_FACTOR'] > 4) & (df['K_FACTOR'] < 30))),
            'sjt13': ((df['DIR_FACTOR'].isna()) | ((df['FACILITY_TYPE'] != 1) | ((df['FACILITY_TYPE'] == 1) & (df['DIR_FACTOR'] == 100)))),
            'sjt14': ((df['DIR_FACTOR'].isna()) | ((df['FACILITY_TYPE'] != 2) | ((df['FACILITY_TYPE'] == 2) & ((df['DIR_FACTOR'] > 50) & (df['DIR_FACTOR'] <= 75))))),
            'sjt15': ((df['FUTURE_AADT'].isna()) | ((df['FUTURE_AADT_VALUE_DATE'].isna()) & (((df['FUTURE_AADT'] > df['AADT']) & (df['FUTURE_AADT'] < (df['AADT'] * 4))) | (df['FUTURE_AADT'] < (df['AADT'] * 0.2)))))
        }

        tmp = df.copy()
        for k,v in spatial_join_checks.items():
            tmp[k] = v
        # tmp2 = pd.DataFrame()
        # for rule in spatial_join_checks.keys():
        #     tmp2 = pd.concat([tmp2, tmp[tmp[rule] == False]])
        return tmp


    def create_output(self, template='cross_validation_rules_template.xlsx', outfilename='cross_validation_summary_old.xlsx'):
        

        isj_errors = self.inventory_spatial_join()
        tsj_errors = self.traffic_spatial_join()

        self.df = isj_errors.merge(tsj_errors, how="outer")
        self.df.columns = [x.upper() for x in self.df.columns.tolist()]
        self.df.rename(columns={"ROUTEID":"RouteID"}, inplace=True)

        
        #Reads sheet on template that list all data items associated with each rule and converts to dictionary
        dataItemsDF = pd.read_excel(template, sheet_name="ruleDataItems", usecols='A,B', nrows=28)
        dataItemsDF['Rule'] = dataItemsDF['Rule'].str.replace("-", "")
        dataItemsDF['Data_Items'].fillna("",inplace=True)
        dataItemsDF['Data_Items'] = dataItemsDF['Data_Items'].str.split(",")
        ruleDict = dict(zip(dataItemsDF['Rule'], dataItemsDF['Data_Items']))
        fwhaRuleDict = ruleDict.copy()
        for rule in ruleDict.keys():
            fwhaRuleDict[rule+"_FHWA"] = fwhaRuleDict.pop(rule)

        for rule in ruleDict.keys():
            try:
                self.df[rule].fillna(True, inplace=True)
            except KeyError:
                pass

        #Reads the rule descripts off of summary sheet and converts to dictionary
        ruleDescDF = pd.read_excel(template, sheet_name="Summary", usecols="A,D")
        ruleDescDF['Rule'] = ruleDescDF['Rule'].str.replace("-", "")
        ruleDesc = dict(zip(ruleDescDF['Rule'], ruleDescDF['Description']))

        #Create copy of template to write to
        shutil.copy(template, outfilename)

        with pd.ExcelWriter(outfilename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            tempDF = self.df.copy()
            numFailed = []
            numPassed = []
            lenFailed = []
            fwhaFailed = []

            #Get counts for failed/passed/length of failed
            print("Updating summary sheet on",outfilename,"...")
            for rule in ruleDict.keys():
                #Assumes all passes for rules not ran
                try:
                    numFailed.append(tempDF[tempDF[rule]==False].shape[0])
                    numPassed.append(tempDF[tempDF[rule]==True].shape[0])
                    lenFailed.append((tempDF[tempDF[rule]==False]['EMP'] - tempDF[tempDF[rule]==False]['BMP']).sum())
                except KeyError:
                    numFailed.append(0)
                    numPassed.append(self.df.shape[0])
                    lenFailed.append(0)

            for rule in fwhaRuleDict.keys():
                try:
                    fwhaFailed.append(tempDF[tempDF[rule]==False].shape[0])
                except KeyError:
                    fwhaFailed.append(0)

            failedDF = pd.DataFrame(numFailed)
            passedDF = pd.DataFrame(numPassed)
            lengthDF = pd.DataFrame(lenFailed)
            fwhaFailedDF = pd.DataFrame(fwhaFailed)

            #Write counts to Summary sheet
            failedDF.to_excel(writer, sheet_name='Summary', startcol=4, startrow=1, header=False, index=False)
            passedDF.to_excel(writer, sheet_name='Summary', startcol=5, startrow=1, header=False, index=False)
            lengthDF.to_excel(writer, sheet_name='Summary', startcol=6, startrow=1, header=False, index=False)
            fwhaFailedDF.to_excel(writer, sheet_name='Summary', startcol=7, startrow=1, header=False, index=False)

            #Create sheets for each rule containing all failed rows (using only columns that the specific rule references)
            for rule in ruleDict.keys():
                tempDF = self.df.copy()
                #Checks to make sure rule has data items associated with it (will be a list if dataItems exists, otherwise will be float (np.nan))
                if type(ruleDict[rule])==list:
                    #Tries using RULENAME (i.e. SJF01) in dataset which is added if the rule is ran
                    #If rule is not run, no column will exist with the rulename, catches KeyError and prints message.
                    try:
                        if tempDF[tempDF[rule]==False].shape[0] > 0:
                            print("Creating error sheet for rule:",rule)
                            dataItems = ['RouteID', 'BMP', 'EMP']
                            [dataItems.append(x) for x in ruleDict[rule] if x not in dataItems]
                            tempDF = tempDF[tempDF[rule]==False]
                            tempDF = tempDF[dataItems]
                            tempDF.to_excel(writer, sheet_name=rule, startrow=1, index=False)
                            worksheet = writer.sheets[rule]
                            worksheet['A1'] = f'=HYPERLINK("#Summary!A1", "Summary Worksheet")'
                            worksheet['A1'].font = Font(underline='single', color='0000EE')
                            #Autofit columns
                            for i in range(1, worksheet.max_column+1):
                                worksheet.column_dimensions[get_column_letter(i)].width = 20
                            #Add rule description to sheet
                            worksheet['B1'] = ruleDesc[rule]

                        else:
                            print("No failed rows for rule:",rule)

                    except KeyError:
                        print(rule,"not found in DF. Sheet was not created in rules_summary.xlsx")
                else:
                    print("No data items for rule",rule,", Sheet not created.")

input_file = 'all_submission_data.csv'
data = pd.read_csv(input_file)

# data = load_defaults(data)

# # Temporary Arnold dataframe until the real one is added
# arnold = pd.DataFrame()
# arnold['RouteID'] = data['RouteID']

# inventory = inventory_spatial_join(data)
# traffic = traffic_spatial_join(data)

# print(inventory)
# print(traffic)



cross_validation = Cross_Validation(data)
cross_validation.create_output()
# isj_errors = cross_validation.inventory_spatial_join()
# tsj_errors = cross_validation.traffic_spatial_join()

# all_errors = isj_errors.merge(tsj_errors, how="outer")


# print('ISJ Errors', isj_errors)
# print('TSJ Errors', tsj_errors)